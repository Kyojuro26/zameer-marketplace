#!/usr/bin/env python3
"""
Unrivaled CRM — normalization pipeline (Phase 1).

Reads the legacy Sales Tracker workbook and produces the clean, normalized store:
Company (primary unit) -> Contact / Project -> Shipment, plus Vendor. Deterministic
and re-runnable. Anything ambiguous is flagged in needs_review.json, never dropped.

Usage:
    python normalize.py --workbook "US-Sales Tracker-2026 .xlsx" --out ./store

Nothing is hard-coded to the operator's data: the workbook path and output dir are
parameters, and companies/contacts/vendors are read from the sheet, not baked in.
"""

import argparse
import json
import os
import pathlib
import re
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run:  pip install openpyxl")


# ---------------------------------------------------------------- helpers

REP_TAG_RE = re.compile(r"\s*\(([^)]{1,12})\)\s*$")   # trailing "(D)", "(Ervin)"
INV_RE = re.compile(r"INV[\s#-]*(\d+)", re.IGNORECASE)
PCT_RE = re.compile(r"(\d{1,3})\s*%")
# A commission/rep-split rate mentioned in the same cell as a "NN%" number is
# NOT a percent-of-invoice-paid figure -- e.g. "Paid, 10% comm to D" reads a
# payment percentage out of a commission rate if PCT_RE is trusted blindly.
# Client feedback 2026-07: this was misreading commission % as percent
# invoiced. Never guess which the number means -- flag it for review instead
# of assigning it as collection/payment status.
# A percentage sitting next to a commission / rep-split mention is a payout
# rate, NOT percent-of-invoice-collected. The original guard was
# r"\bcomm(?:ission)?\b", which missed the plural: "10% commissions to D" was
# stored as partial:10%. That is the exact client-reported incident this guard
# was written for, reachable by adding one letter -- so it is spelled out
# generously here, and a miss costs only a needs_review entry.
COMMISSION_RE = re.compile(
    r"\b(?:comm|comms|commn|comm'?n|cmsn|commis(?:s)?ion(?:s)?|"
    r"commision(?:s)?|rep\s*split|split|payout|kickback)\b\.?",
    re.IGNORECASE)
# "10% to D" / "15% for JS" -- a percentage allocated TO someone is a person's
# share, not a collection status. Requires a name-shaped token so "50% to be
# invoiced" and "10% for freight" do not trip it.
REP_ALLOC_RE = re.compile(r"%\s*(?:to|for)\s+(?:[A-Z]{1,3}\b|[A-Z][a-z]+)")


def commission_like(text):
    """True when a percentage in this text is a payout rate, not a collection
    percentage. Both call sites treat True the same way: never guess a status
    from the number, leave it open, and flag it for a person."""
    t = "" if text is None else str(text)
    return bool(COMMISSION_RE.search(t)) or bool(REP_ALLOC_RE.search(t))
# Payment wording lives in ONE place -- see payment_words.py. Three modules
# need it and a checker that drifts from the importer cannot catch the
# importer's mistake, which is exactly how "NOT PAID" read as paid for five
# releases. Imported by path so these stay runnable as plain scripts.
def _load_payment_words():
    import importlib.util, os
    _p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "payment_words.py")
    _spec = importlib.util.spec_from_file_location("crm_payment_words", _p)
    _m = importlib.util.module_from_spec(_spec)
    _spec.loader.exec_module(_m)
    return _m


_PW = _load_payment_words()
PAID_RE = _PW.PAID_RE
says_paid = _PW.says_paid


def st_text(v):
    return "" if v is None else str(v)
LEADING_NUMS_RE = re.compile(r"^\s*(\d{2,6})(?:\s*(?:and|&|,|/)\s*(\d{2,6}))*")
ONE_NUM_RE = re.compile(r"\d{2,6}")



# Row bounds exist so a sheet with a stray cell far down does not make
# openpyxl walk a million empty rows. They must never truncate real data
# silently: the ledger grows past them on its own, and the independent audit
# re-read with the SAME numbers, so it could not catch it either.
ROW_CAP = 20000


def _cap_check(ws, cap, label, review, min_row=1):
    """Flag loudly if the sheet holds data beyond the row bound."""
    try:
        beyond = ws.max_row or 0
    except Exception:                                 # noqa: BLE001
        return
    if beyond > cap:
        review.append({
            "type": "sheet_truncated",
            "detail": (f"{label}: the sheet has {beyond} rows but only the first "
                       f"{cap} were read. Rows {cap + 1}-{beyond} were NOT "
                       f"imported. Raise ROW_CAP in normalize.py and re-run."),
            "sheet_row": cap,
        })


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# A parenthetical counts as an OWNER (rep) only if it looks like initials:
# 1-3 letters, or "X & Y" of initials. Everything else is a note, NOT a rep.
#
# Client feedback 2026-07-07: "the text in parentheses across records are just notes" and
# a customer must be ONE record regardless of the note that trails its name.
# So the company identity is the text BEFORE the first '('; every parenthetical
# group and any trailing free-note is pulled off as owner-tag or note and never
# becomes part of the company name/key.
OWNER_TAG_RE = re.compile(r"^[A-Za-z]{1,3}(\s*&\s*[A-Za-z]{1,3})?$")
CONFIRMED_REPS = {"D", "G"}
PAREN_RE = re.compile(r"\(([^()]*)\)")

# Known misspellings that should collapse to one company (explicit + reviewable,
# so distinct firms are never merged by guesswork — only these exact typos).
SPELLING_FIXES = {
    "fullfillment": "fulfillment", "fulfilment": "fulfillment",
    "fullfilment": "fulfillment", "fullfiment": "fulfillment",
    "incorportated": "incorporated", "oringinal": "original",
}

# base names that are really notes/labels, not customers
NOTE_WORDS = {"stay", "follow", "waiting", "just", "got", "need", "needs", "watch",
              "working", "sent", "told", "tentatively", "should", "confirm", "action"}


def strip_rep_tags(name):
    """Return (clean_name, [owners], [notes]).

    clean_name is the text before the first '('. Parenthetical groups are
    classified as owner initials (D, G, "D & G") or freeform notes; any prose
    trailing the name (e.g. "-Just got a verbal...") is captured as a note too.
    """
    owners, notes = [], []
    if not name:
        return None, owners, notes
    s = str(name).strip()
    cut = s.find("(")
    base = (s[:cut] if cut != -1 else s).strip()
    for m in PAREN_RE.finditer(s):
        tag = m.group(1).strip()
        if not tag:
            continue        # skip-ok: empty () in a name, not a data row
        (owners if OWNER_TAG_RE.match(tag) else notes).append(tag)
    if cut != -1:
        leftover = re.sub(r"\s+", " ", PAREN_RE.sub(" ", s[cut:])).strip(" -–—")
        if leftover:
            notes.append(leftover)
    # de-dupe, preserve order
    owners = list(dict.fromkeys(owners))
    notes = list(dict.fromkeys(notes))
    return (base or None), owners, notes


def _is_note_base(b):
    """True when a base name is a note/label fragment, not a company name."""
    if not b:
        return True
    if b[0] in "-–—" or "/" in b:
        return True
    low = b.lower()
    if low in PLACEHOLDERS:
        return True
    words = low.split()
    return len(words) >= 3 and any(w.strip(".,") in NOTE_WORDS for w in words)


def looks_like_date(s):
    s = str(s).strip()
    return bool(re.match(r"^\d{4}-\d\d-\d\d", s) or re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", s))


def looks_like_po(s):
    s = str(s).strip().lower()
    return s.startswith("po ") or s.startswith("po#") or "po #" in s or bool(re.match(r"^p\d", s))


PLACEHOLDERS = {"tbd", "na", "n/a", "none", "pending", "?", "-"}
# words that mark a cell as a status/note fragment, not a client name
STATUS_WORDS = {"paid", "prepay", "prepaid", "deposit", "balance", "due", "signed",
                "consent", "quote", "order", "invoice", "shipped", "delivered",
                "install", "installed", "hold", "complete", "completed", "credit",
                "refund", "net", "rma", "email", "customer", "date"}


def is_nameish(s):
    """A plausible company/client name: not a date, PO, address, placeholder, status,
    order/invoice number, or note fragment."""
    s = clean(s)
    if not s:
        return False
    base, _, _ = strip_rep_tags(s)          # ignore trailing "(D)" when judging
    base = (base or "").strip()
    if not base or base.lower() in PLACEHOLDERS:
        return False
    if looks_like_date(base) or looks_like_po(base):
        return False
    if re.search(r"[\d/:#]", base):          # numbers, order#s, dates, "x/y" notes
        return False
    if any(w in STATUS_WORDS for w in re.split(r"[^a-z]+", base.lower()) if w):
        return False
    if len(re.sub(r"[^A-Za-z]", "", base)) < 2 or len(base) > 45:
        return False
    return True


def pick_client(cells):
    """From a Project Tracker row's middle cells, pick the first plausible client name.
    The columns drift row-to-row, so scan rather than trust a fixed position."""
    for v in cells:
        if is_nameish(v):
            return clean(strip_rep_tags(clean(v))[0])
    return None


def slugify_company(name):
    """Normalized company key: name before the first '(', trailing punctuation and
    notes removed, collapsed and casefolded. Rejects date/PO/number-only and
    note-fragment cells so those never become companies. Common misspellings are
    collapsed so obvious typo-variants merge to one key."""
    base, _, _ = strip_rep_tags(name)
    if not base:
        return None
    base = re.sub(r"[.,;:]+$", "", base).strip()          # trailing punctuation
    base = re.sub(r"\s+", " ", base)                        # collapse whitespace
    if looks_like_date(base) or looks_like_po(base):
        return None
    if _is_note_base(base):                                # "-Stay on...", "Follow up/..."
        return None
    if re.fullmatch(r"[\d/.\-\s]+", base):                 # pure number/date-ish
        return None
    if len(re.sub(r"[^A-Za-z]", "", base)) < 2:            # need >=2 letters
        return None
    slug = re.sub(r"[^a-z0-9]+", "-", base.lower()).strip("-")
    slug = "-".join(SPELLING_FIXES.get(t, t) for t in slug.split("-"))
    return slug or None


def display_company(name):
    base, _, _ = strip_rep_tags(name)
    if not base:
        return None
    return re.sub(r"\s+", " ", re.sub(r"[.,;:]+$", "", base).strip())


def domain_of(email):
    e = clean(email)
    if e and "@" in e:
        return e.split("@")[-1].strip().lower()
    return None


def parse_project_key(raw):
    """
    Parse a Project Tracker key cell into {project_nos, invoice_no, collection_status,
    review}. Handles '4521 (INV 9001-PAID)', '4522 (INV 9002-50%)', '4530 and 4531',
    'Word Proposal', 'Check'.
    """
    out = {"project_nos": [], "invoice_no": None, "collection_status": None,
           "review": None, "raw": raw}
    if raw is None:
        out["review"] = "empty key"
        return out
    s = str(raw).strip()

    m = LEADING_NUMS_RE.match(s)
    if m:
        # capture every 2-6 digit run in the leading segment (handles "4530 and 4531")
        lead = m.group(0)
        out["project_nos"] = ONE_NUM_RE.findall(lead)
    if not out["project_nos"]:
        out["review"] = f"non-numeric key: {s!r}"
        return out
    if len(out["project_nos"]) > 1:
        out["review"] = f"multi-project row: {s!r}"

    inv = INV_RE.search(s)
    if inv:
        out["invoice_no"] = inv.group(1)

    _paid = says_paid(s)
    if _paid is True:
        out["collection_status"] = "paid"
    elif _paid is None:
        # a qualified "paid" -- "not paid", "will be paid", "partially paid".
        # Never guess a collection status from it; flag for a person.
        out["review"] = (out["review"] + "; " if out["review"] else "") + \
            f"payment wording is qualified, collection_status left unset: {s!r}"
        if inv:
            out["collection_status"] = "open"
    elif commission_like(s):
        # A percentage here is a commission/rep-split rate, not percent paid --
        # never guess; flag it so a person confirms the real collection status.
        out["review"] = (out["review"] + "; " if out["review"] else "") + \
            f"commission % present, collection_status left unset: {s!r}"
        if inv:
            out["collection_status"] = "open"
    else:
        pct = PCT_RE.search(s)
        if pct:
            n = int(pct.group(1))
            out["collection_status"] = "paid" if n >= 100 else f"partial:{n}%"
        elif inv:
            out["collection_status"] = "open"   # invoiced, no status marker
    return out


def num(v):
    if isinstance(v, (int, float)):
        return v
    if v is None:
        return None
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def header_map(ws, header_row, ncols):
    m = {}
    for c in range(1, ncols + 1):
        h = clean(ws.cell(row=header_row, column=c).value)
        if h:
            m[re.sub(r"[\s:]+$", "", h).lower()] = c
    return m


def col(hm, *names):
    for n in names:
        key = n.lower()
        if key in hm:
            return hm[key]
    return None


# ---------------------------------------------------------------- pipeline

def _guard_live_store(outdir, force, mode="replace"):
    """Refuse to overwrite a store that already holds user data / edits.

    normalize.py rewrites every entity file wholesale; pointed at the live
    production store it would destroy every edit made since migration. A
    non-empty changelog.jsonl (edits are logged there) or an already-populated
    companies.json is the signal that this is a real store, not an empty output
    directory. A brand-new store whose files are just "[]" (2 bytes) does not
    trip this. Pass force=True only for a throwaway/scratch target. (v0.1.14)

    mode="merge" is exempt because it does not overwrite wholesale -- see
    merge.py, which refuses on its own terms if it cannot preserve operator
    work. The guard exists to stop a REPLACE landing on a live store.
    """
    if force or mode == "merge":
        return
    signals = []
    changelog = os.path.join(outdir, "changelog.jsonl")
    companies = os.path.join(outdir, "companies.json")
    if os.path.exists(changelog) and os.path.getsize(changelog) > 0:
        signals.append("changelog.jsonl (edit history present)")
    if os.path.exists(companies) and os.path.getsize(companies) > 2:
        signals.append("companies.json (existing records present)")
    if signals:
        raise SystemExit(
            f"REFUSING to write into what looks like a LIVE store at {outdir!r}:\n"
            "  - " + "\n  - ".join(signals) + "\n\n"
            "normalize.py rewrites every entity file wholesale and would destroy "
            "any edits made since migration. If you are certain this directory is "
            "a throwaway/scratch target, re-run with --force.")


def run(workbook, outdir, force=False, mode="merge"):
    _guard_live_store(outdir, force, mode)
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    companies = OrderedDict()     # company_id -> record
    contacts = []
    projects = OrderedDict()      # project_no -> record
    shipments = []
    invoices = []                 # CLIENT Invoices table = receivables ledger
    vendors = OrderedDict()
    review = []

    def ensure_company(name, role="customer"):
        cid = slugify_company(name)
        if not cid:
            return None
        if cid not in companies:
            companies[cid] = {"company_id": cid, "display_name": display_company(name),
                              "role": role, "domains": [], "locations": []}
        elif role == "vendor":
            companies[cid]["role"] = "vendor"
        return cid

    # ---- Sales Tracker sheets (deals -> projects) ----
    for sheet in [s for s in wb.sheetnames if s.lower().startswith("sales tracker")]:
        ws = wb[sheet]
        year = 2025 if "2025" in sheet else (2026 if "2026" in sheet else None)
        # find header row (cell A == "Project#")
        hrow = None
        for rr in range(1, 8):
            if clean(ws.cell(row=rr, column=1).value) == "Project#":
                hrow = rr
                break
        if not hrow:
            review.append({
                "type": "sheet_header_not_found",
                "detail": (f"{ws.title!r}: no 'Project#' header in the first 7 "
                           f"rows, so NONE of this sheet was imported -- no "
                           f"deals, revenue, margin or collection status. Check "
                           f"the header spelling (a stray space is enough)."),
            })
            continue
        hm = header_map(ws, hrow, 15)
        c_pno = col(hm, "project#")
        c_date = col(hm, "date")
        c_cust = col(hm, "customer")
        c_desc = col(hm, "description")
        c_loc = col(hm, "location")
        c_stat = col(hm, "status")
        c_poyn = col(hm, "po y/n")
        c_cpo = col(hm, "client po#", "po#")
        c_inv = col(hm, "invoice #", "invoice#")
        c_rev = col(hm, "revenue")
        c_cost = col(hm, "total cost")
        c_gp = col(hm, "total gp")
        c_marg = col(hm, "margain", "margin")
        c_notes = col(hm, "notes")

        for row in ws.iter_rows(min_row=hrow + 1, max_row=ROW_CAP, values_only=True):
            def g(ci):
                return row[ci - 1] if ci and ci - 1 < len(row) else None
            cust = clean(g(c_cust))
            pno_raw = g(c_pno)
            if not cust and pno_raw in (None, ""):
                if any(clean(g(c)) for c in (c_rev, c_gp, c_cost) if c):
                    review.append({
                        "type": "deal_row_without_keys",
                        "detail": ("row carries financial values but has neither "
                                   "a customer nor a project number, so it was "
                                   "not imported"),
                        "sheet_row": rn,
                    })
                continue
            pno = clean(pno_raw)
            if pno:
                pno = re.sub(r"\.0$", "", pno)
            status_val = (clean(g(c_stat)) or "").lower() or None
            rev_val = num(g(c_rev))
            # A row with a customer cell but no project#, status, or revenue is a
            # stray label/note row. In the source workbook these read like
            # "Legend", a person's name followed by "Action Item", or
            # "Follow up/...". Not a deal — skip it so it never mints a
            # phantom customer.
            if cust and not (pno or status_val or (rev_val not in (None, 0, 0.0))):
                review.append({"type": "label_row_skipped", "sheet": sheet,
                               "customer": str(cust)[:60]})
                continue
            clean_name, owners, annos = strip_rep_tags(cust) if cust else (None, [], [])
            cid = ensure_company(cust, "customer") if cust else None
            loc = clean(g(c_loc))
            if cid and loc:
                companies[cid]["locations"].append(loc)
            rec = {
                "project_no": pno,
                "company_id": cid,
                "company_name": clean_name,
                "owner": owners,
                "date": str(g(c_date)) if g(c_date) else None,
                "description": clean(g(c_desc)),
                "location": loc,
                "annotations": annos,
                "status": (clean(g(c_stat)) or "").lower() or None,
                "po_flag": (str(g(c_poyn)).strip().lower() == "yes") if g(c_poyn) else None,
                "client_po_no": clean(g(c_cpo)),
                "invoice_no": re.sub(r"\.0$", "", clean(g(c_inv))) if clean(g(c_inv)) else None,
                "collection_status": None,
                "revenue": num(g(c_rev)),
                "total_cost": num(g(c_cost)),
                "gross_profit": num(g(c_gp)),
                "margin": num(g(c_marg)),
                "notes": clean(g(c_notes)),
                "year": year,
            }
            # flag financial cells that silently failed to parse (e.g. margin "P")
            for label, ci in (("revenue", c_rev), ("total_cost", c_cost),
                              ("gross_profit", c_gp), ("margin", c_marg)):
                raw = g(ci)
                if raw is not None and num(raw) is None:
                    review.append({"type": "non_numeric_financial", "field": label,
                                   "project_no": pno, "sheet": sheet,
                                   "value": str(raw)[:40]})
            if pno and pno in projects:
                # Duplicate project number in the source workbook (same number
                # reused across sheets/years). Keep the FIRST occurrence under the
                # bare `pno` so existing shipment/invoice references still resolve,
                # and store later occurrences under a distinct, addressable id
                # (project_no rewritten to match) so update_project/get_project can
                # target each one instead of two records silently colliding on the
                # same identity. (v0.1.14)
                dup_base = f"{pno}#{year}" if year else f"{pno}#dup"
                dup_id, _n = dup_base, 2
                while dup_id in projects:   # 3+ reuses in the same year: stay unique
                    dup_id, _n = f"{dup_base}-{_n}", _n + 1
                rec["project_no"] = dup_id
                review.append({"type": "duplicate_project_no", "project_no": pno,
                               "stored_as": dup_id, "sheet": sheet,
                               "detail": ("Reuse of a project number from an earlier "
                                          "row; stored under a distinct id so both are "
                                          "individually editable. Confirm which is current.")})
                projects[dup_id] = rec
            elif pno:
                projects[pno] = rec
            else:
                review.append({"type": "project_without_number", "customer": cust, "sheet": sheet})
                projects[f"noid-{len(projects)}"] = rec

    # ---- Client Contacts (people -> contacts + companies) ----
    ws = wb["Client Contacts"]
    hm = header_map(ws, 1, 26)
    cc = {k: col(hm, k) for k in ["client business", "client name", "email",
          "phone number", "job title", "location", "action taken and notes",
          "last date of action"]}
    _cap_check(ws, ROW_CAP, 'contacts sheet', review)
    for row in ws.iter_rows(min_row=2, max_row=ROW_CAP, values_only=True):
        def g(ci):
            return row[ci - 1] if ci and ci - 1 < len(row) else None
        biz = clean(g(cc["client business"]))
        name = clean(g(cc["client name"]))
        if not biz and not name:
            continue        # skip-ok: a contact row with neither a business nor a name carries nothing to import
        cid = ensure_company(biz, "customer") if biz else None
        email = clean(g(cc["email"]))
        d = domain_of(email)
        if cid and d and d not in companies[cid]["domains"]:
            companies[cid]["domains"].append(d)
        contacts.append({
            "company_id": cid, "company_name": display_company(biz) if biz else None,
            "name": name, "email": email, "phone": clean(g(cc["phone number"])),
            "title": clean(g(cc["job title"])), "location": clean(g(cc["location"])),
            "action_notes": clean(g(cc["action taken and notes"])),
            "last_action": str(g(cc["last date of action"])) if g(cc["last date of action"]) else None,
        })

    # ---- Vendor Contacts ----
    ws = wb["Vendor Contacts"]
    vhrow = None
    for rr in range(1, 5):
        if clean(ws.cell(row=rr, column=1).value) in ("Company", "Company:"):
            vhrow = rr
            break
    vhrow = vhrow or 2
    hm = header_map(ws, vhrow, 26)
    vc = {k: col(hm, *a) for k, a in {
        "company": ["company"], "hq": ["headquarters location"],
        "rep": ["sales rep/contact"], "email": ["contact email"],
        "phone": ["contact phone number"], "offerings": ["offerings"],
        "send_po": ["send po's to", "send po’s to"], "send_inv": ["send invoices to"],
    }.items()}
    _cap_check(ws, ROW_CAP, 'vendors sheet', review)
    for row in ws.iter_rows(min_row=vhrow + 1, max_row=ROW_CAP, values_only=True):
        def g(ci):
            return row[ci - 1] if ci and ci - 1 < len(row) else None
        comp = clean(g(vc["company"]))
        if not comp:
            continue        # skip-ok: a vendor row with no company name carries nothing to import
        cid = ensure_company(comp, "vendor")
        rep = clean(g(vc["rep"]))
        email = clean(g(vc["email"]))
        vendors[cid] = {
            "company_id": cid, "display_name": display_company(comp),
            "hq_location": clean(g(vc["hq"])), "rep": rep, "email": email,
            "phone": clean(g(vc["phone"])), "offerings": clean(g(vc["offerings"])),
            "po_routing": clean(g(vc["send_po"])), "invoice_routing": clean(g(vc["send_inv"])),
            "po_routing_source": "sheet" if clean(g(vc["send_po"])) else "knowledge-base (to mine)",
        }
        if rep or email:
            contacts.append({"company_id": cid, "company_name": display_company(comp),
                             "name": rep, "email": email, "phone": clean(g(vc["phone"])),
                             "title": "Vendor contact", "location": clean(g(vc["hq"])),
                             "action_notes": None, "last_action": None})

    # ---- Project Tracker: TWO stacked tables (fixed 2026-07-03 after audit) --
    #
    # Rows 2..(invoice header)  OPEN ORDERS: A=Unrivaled Project#, B=Client PO#,
    #   C=Start, D=Client, E=Location, F=Notes, G..T = Vendor N PO#/Ship Date.
    # Row with A=="Invoice Number:" starts the CLIENT INVOICES table:
    #   A=Invoice Number (== customer order #), B=PO#/Order Number (client
    #   PO/contract), C=Invoice Date, D=payment status (header says "DUE Date"
    #   but values are Paid/etc.), E=Client, F=payment notes, G.. = EITHER a
    #   pay date and "Vendors were…" prose OR real vendor-PO shipment legs.
    # Semantics confirmed by Zeeshan 2026-07-03.
    ws = wb["Project Tracker"]
    _cap_check(ws, ROW_CAP, 'project tracker sheet', review)
    all_rows = list(ws.iter_rows(min_row=1, max_row=ROW_CAP, values_only=True))
    inv_header_idx = next((i for i, r in enumerate(all_rows)
                           if r and clean(r[0]) == "Invoice Number:"), None)
    if inv_header_idx is None:
        review.append({"type": "data_quality_note",
                       "detail": "CLIENT Invoices header row not found in Project "
                                 "Tracker; invoice table skipped."})
        inv_header_idx = len(all_rows)

    DATEISH_RE = re.compile(r"^\d{4}-\d\d-\d\d|^\d{1,2}/\d{1,2}/\d{2,4}$")
    PO_TOKEN_RE = re.compile(r"(?i)(\bp\.?o\.?\s*#?\s*\d|\bpo\b|^p\d{4,})")

    def _is_dateish(v):
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return True
        s = clean(v)
        return bool(s and DATEISH_RE.match(s))

    def _is_po_token(v):
        s = clean(v)
        return bool(s and not _is_dateish(v) and PO_TOKEN_RE.search(s))

    # A running counter PER sid_base, not per row. It used to restart at 1 on
    # every row, so "4521" appearing on two open-order rows (phase 2, a
    # multi-project key, a continuation) minted 4521-L1 twice. Duplicate
    # shipment_ids are unusable downstream: update_shipment and
    # reassign_shipment cannot tell the legs apart and refuse outright, and the
    # visual app's .find() silently opens the first one's data.
    _leg_seq = {}

    def _next_leg(sid_base):
        _leg_seq[sid_base] = _leg_seq.get(sid_base, 0) + 1
        return _leg_seq[sid_base]

    def add_shipment(sid_base, leg_no, po_val, sd_val, pnos, company_id,
                     client, notes, start, invoice_no=None):
        stage = "Shipped" if sd_val else "Ordered"
        if re.search(r"\bhold\b", notes or "", re.IGNORECASE):
            stage = "On Hold"
        primary = pnos[0] if pnos else None
        shipments.append({
            # leg_no is ignored in favour of the running counter
            "shipment_id": f"{sid_base}-L{_next_leg(sid_base)}",
            "project_no": primary, "all_project_nos": [p for p in pnos if p],
            "invoice_no": invoice_no,
            "vendor_po_raw": po_val,
            "ship_date": str(sd_val) if sd_val else None,
            "stage": stage,
            "company_id": company_id,
            "client_name": client,
            "linked_to_project": bool(primary and primary in projects),
            "open_orders_notes": notes,
            "start_date": str(start) if start else None,
        })

    # --- Table 1: open orders ---
    for r in all_rows[1:inv_header_idx]:
        if not r or r[0] is None:
            if r and any(c is not None and str(c).strip() for c in r[1:]):
                review.append({
                    "type": "open_order_row_without_project",
                    "detail": ("row has content but no project number, so its "
                               "shipment legs were not imported: "
                               + repr([clean(c) for c in list(r)[:6]])),
                })
            continue
        raw_key = r[0]
        parsed = parse_project_key(raw_key)
        cells = list(r) + [None] * 24
        picked_client = pick_client([cells[1], cells[2], cells[3], cells[4]])
        if parsed["review"]:
            review.append({"type": "project_key", "detail": parsed["review"],
                           "client": picked_client, "raw": str(raw_key)})
        pnos = parsed["project_nos"] or [None]
        primary = pnos[0]
        row_company = projects[primary]["company_id"] if primary in projects else None
        if not row_company and picked_client:
            row_company = ensure_company(picked_client, "customer")
        for pno in pnos:
            if pno and pno in projects:
                if parsed["invoice_no"] and not projects[pno]["invoice_no"]:
                    projects[pno]["invoice_no"] = parsed["invoice_no"]
                if parsed["collection_status"]:
                    projects[pno]["collection_status"] = parsed["collection_status"]
        leg = 0
        for j in range(6, 20, 2):   # G..T as PO/date pairs
            po_val, sd_val = clean(cells[j]), cells[j + 1]
            if not po_val and not sd_val:
                continue    # skip-ok: an empty PO/date column pair, not a row
            leg += 1
            add_shipment(primary or "noid", leg, po_val, sd_val, pnos,
                         row_company, picked_client, clean(cells[5]), cells[2])

    # --- Table 2: client invoices (the receivables ledger) ---
    # index deal-log invoice numbers for project linking
    proj_by_inv = {}
    for pno, p in projects.items():
        if p.get("invoice_no"):
            proj_by_inv.setdefault(re.sub(r"\.0$", "", str(p["invoice_no"])), pno)

    for rn, r in enumerate(all_rows[inv_header_idx + 1:], inv_header_idx + 2):
        if not r or r[0] is None:
            # A blank invoice-number cell used to drop the row entirely, with no
            # needs_review entry -- and that is exactly what a merged cell, or
            # an order not yet invoiced, looks like. A row carrying any other
            # content is a real receivable that would simply never exist.
            if r and any(c is not None and str(c).strip() for c in r[1:]):
                review.append({
                    "type": "invoice_row_without_number",
                    "detail": ("row has content but no invoice number, so it was "
                               "not imported: "
                               + repr([clean(c) for c in list(r)[:6]])),
                    "sheet_row": rn,
                })
            continue
        cells = list(r) + [None] * 24
        invoice_no = clean(r[0])
        if invoice_no:
            invoice_no = re.sub(r"\.0$", "", invoice_no)
        if not invoice_no or not re.fullmatch(r"\d{2,6}", invoice_no):
            review.append({"type": "invoice_key", "detail": f"non-numeric invoice "
                           f"number: {invoice_no!r}", "sheet_row": rn})
        client_raw = clean(cells[4])
        client_name, _, _ = strip_rep_tags(client_raw) if client_raw else (None, [], [])
        company_id = ensure_company(client_raw, "customer") if client_raw else None
        status_raw = clean(cells[3])
        pay_notes = clean(cells[5])
        blob = f"{status_raw or ''} {pay_notes or ''}"
        _paid = says_paid(blob)
        if _paid is True:
            payment_status = "paid"
        elif _paid is None:
            # qualified wording: leave it OPEN (the safe direction for a
            # receivable -- an open invoice gets chased, a wrongly-paid one
            # never does) and flag it.
            payment_status = "open"
            review.append({"type": "payment_wording_ambiguous", "invoice_no": invoice_no,
                           "detail": f"qualified payment wording, not read as paid: {blob!r}",
                           "sheet_row": rn})
        elif commission_like(blob):
            # Same guard as parse_project_key: a "NN%" next to a commission
            # mention is a rep-split rate, not percent of the invoice paid --
            # never guess. Leave it open and flag it for a person to confirm.
            payment_status = "open"
            review.append({"type": "commission_percent_ignored", "invoice_no": invoice_no,
                           "detail": f"commission % present, not read as payment_status: {blob!r}",
                           "sheet_row": rn})
        else:
            pct = PCT_RE.search(blob)
            payment_status = f"partial:{pct.group(1)}%" if pct else "open"

        # walk the tail columns: pay date vs real vendor-PO legs vs prose
        legs, pay_date, vendor_notes, cur_po = [], None, [], None
        for j in range(6, 24):
            v = cells[j]
            if v is None:
                continue
            if _is_po_token(v):
                if cur_po:
                    legs.append((cur_po, None))
                cur_po = clean(v)
            elif _is_dateish(v):
                if cur_po:
                    legs.append((cur_po, v))
                    cur_po = None
                elif pay_date is None:
                    pay_date = v
            else:
                vendor_notes.append(clean(v))
        if cur_po:
            legs.append((cur_po, None))

        # link invoice -> project (deal-log Invoice # column, else key-parse
        # enrichment done above); never guess by name
        project_no = proj_by_inv.get(invoice_no)
        if project_no and payment_status:
            if not projects[project_no].get("collection_status"):
                projects[project_no]["collection_status"] = payment_status

        invoices.append({
            "invoice_no": invoice_no,
            "client_po_raw": clean(cells[1]),
            "invoice_date": str(cells[2]) if cells[2] else None,
            "payment_status": payment_status,
            "payment_status_raw": status_raw,
            "pay_date": str(pay_date) if pay_date else None,
            "client_name": client_name,
            "company_id": company_id,
            "payment_notes": pay_notes,
            "vendor_notes": [v for v in vendor_notes if v],
            "project_no": project_no,
            "sheet_row": rn,
        })
        pnos = [project_no] if project_no else []
        for leg_no, (po_val, sd_val) in enumerate(legs, 1):
            add_shipment(project_no or f"inv{invoice_no}", leg_no, po_val, sd_val,
                         pnos, company_id, client_name, pay_notes, None,
                         invoice_no=invoice_no)

    # ---- finalize companies (dedupe locations/domains) ----
    for c in companies.values():
        c["locations"] = list(dict.fromkeys(c["locations"]))[:5]
        c["primary_location"] = c["locations"][0] if c["locations"] else None

    # flag distinct owner tags that aren't confirmed reps -> for the operator's legend
    all_owner_tags = set()
    for p in projects.values():
        all_owner_tags.update(p.get("owner", []))
    unconfirmed = sorted(t for t in all_owner_tags if t not in CONFIRMED_REPS)
    if unconfirmed:
        review.append({"type": "owner_tags_pending_legend",
                       "detail": "confirm which are reps vs. descriptors",
                       "tags": unconfirmed})

    # flag possible duplicate companies (same first two significant tokens)
    STOP = {"inc", "llc", "co", "corp", "the", "and", "of"}
    buckets = {}
    for c in companies.values():
        if c["role"] == "vendor":
            continue
        toks = [t for t in re.split(r"[^a-z0-9]+", (c["display_name"] or "").lower())
                if t and t not in STOP]
        if not toks:
            continue        # skip-ok: token loop for duplicate detection, not a row
        key = " ".join(toks[:2]) if len(toks) >= 2 else toks[0]
        buckets.setdefault(key, []).append(c["company_id"])
    dupes = {k: v for k, v in buckets.items() if len(v) > 1}
    for k, ids in sorted(dupes.items()):
        review.append({"type": "possible_duplicate_companies", "key": k, "company_ids": ids})

    # shipments we couldn't attach to a company (no clean client name in the row)
    orphan_ships = [s["shipment_id"] for s in shipments if not s["company_id"]]
    if orphan_ships:
        review.append({"type": "shipments_without_company", "count": len(orphan_ships),
                       "shipment_ids": orphan_ships})

    # data-quality note: how well the two tabs join on project number
    linked = sum(1 for s in shipments if s["linked_to_project"])
    review.append({"type": "data_quality_note",
                   "detail": (f"Only {linked}/{len(shipments)} shipments share a project# "
                              "with the deal log; the rest are attached by client name. "
                              "Open-orders and deal-log project numbers largely don't overlap "
                              "in the source workbook — a known migration reality, not a bug.")})

    os.makedirs(outdir, exist_ok=True)
    # flag orphans (no company) explicitly — never silently dropped
    for p in projects.values():
        if not p.get("company_id"):
            review.append({"type": "project_without_company",
                           "project_no": p.get("project_no"),
                           "raw_company_name": p.get("company_name")})
    for c in contacts:
        if not c.get("company_id"):
            review.append({"type": "contact_without_company",
                           "contact_name": c.get("name")})

    out = {
        "companies.json": list(companies.values()),
        "contacts.json": contacts,
        "projects.json": list(projects.values()),
        "shipments.json": shipments,
        "invoices.json": invoices,
        "vendors.json": list(vendors.values()),
        "needs_review.json": review,
    }
    # Take the store's OWN write lock around the read-merge-write. Without it
    # the importer read the store, spent seconds merging, then replaced every
    # file -- and any edit a tool made in that window was discarded, with
    # ok:true reported to both sides. That is precisely the failure
    # Store.write_lock exists for, and the importer was the one writer not
    # using it. Best-effort: if the lock cannot be loaded (running this file
    # standalone, outside the plugin), carry on rather than refuse to import.
    _lock = _store_write_lock(outdir)
    with _lock:
        return _finish(out, outdir, mode, workbook, companies, contacts,
                       projects, shipments, invoices, vendors, review)


def _store_write_lock(outdir):
    """server.Store's write lock, taken WITHOUT constructing a Store.

    Store.__init__ has side effects -- it creates missing entity files, writes
    .store-manifest.json and sweeps temp files -- so merely asking for a lock
    used to mutate the store before the import had decided anything, and
    populate a full empty store in a mistyped output directory.

    Worse, it can RAISE: the missing-file guard refuses to open a store whose
    changelog proves a file existed. Wrapped in a bare `except Exception ->
    nullcontext`, that refusal was swallowed and the import then ran with NO
    LOCK AT ALL -- silently recreating the race this function exists to close,
    while bypassing a data-loss guard. __new__ takes the lock with none of
    that: write_lock only ever reads self.root.
    """
    import contextlib
    import importlib.util
    import os as _os
    p = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                      "..", "mcp", "server.py")
    try:
        spec = importlib.util.spec_from_file_location("crm_server_for_lock", p)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
    except Exception as e:                            # noqa: BLE001
        # genuinely no server module (running this file outside the plugin).
        # Say so -- a silently unlocked import is how concurrent edits vanish.
        print(f"WARNING: could not load the store lock from {p} ({e}). "
              f"Proceeding WITHOUT a lock -- close the CRM before importing.",
              file=sys.stderr)
        return contextlib.nullcontext()
    st = mod.Store.__new__(mod.Store)          # no __init__: no side effects
    st.root = pathlib.Path(outdir)
    return st.write_lock()


def _finish(out, outdir, mode, workbook, companies, contacts, projects,
            shipments, invoices, vendors, review):
    merge_report = None
    if mode == "merge":
        # Default. Refresh from the workbook without discarding operator work;
        # merge.merge_all raises if it cannot do that safely.
        try:
            from . import merge as _merge_mod   # packaged
        except ImportError:
            import merge as _merge_mod          # run as a script
        out, merge_report = _merge_mod.merge_all(out, outdir)
    # Stage every file before replacing any: a failure partway through used to
    # leave companies/contacts/projects NEW and invoices/shipments OLD, cross-
    # referencing keys that no longer agree -- and Store then auto-created the
    # missing ones empty and booted reporting zero invoices.
    staged = []
    try:
        for fname, data in out.items():
            tmp = os.path.join(outdir, f".~{fname}.tmp")
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, default=str)
                f.flush()
                os.fsync(f.fileno())
            staged.append((tmp, os.path.join(outdir, fname)))
        for tmp, target in staged:
            os.replace(tmp, target)
    except BaseException:
        for tmp, _ in staged:
            try:
                if os.path.exists(tmp):
                    os.unlink(tmp)
            except OSError:
                pass
        raise

    summary = {
        "companies": len(companies),
        "  of which vendors": sum(1 for c in companies.values() if c["role"] == "vendor"),
        "contacts": len(contacts),
        "projects": len(projects),
        "  won": sum(1 for p in projects.values() if p["status"] == "won"),
        "  pending": sum(1 for p in projects.values() if p["status"] == "pending"),
        "  with collection_status": sum(1 for p in projects.values() if p["collection_status"]),
        "shipments": len(shipments),
        "invoices": len(invoices),
        "  invoices paid": sum(1 for i in invoices if i["payment_status"] == "paid"),
        "  invoices linked to a project": sum(1 for i in invoices if i["project_no"]),
        "vendors": len(vendors),
        "needs_review": len(review),
        "merge_report": merge_report,
    }
    return summary


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--out", default="./store")
    ap.add_argument("--replace", action="store_true",
                    help="REPLACE the store wholesale instead of merging. "
                         "Discards every hand-entered record and every edit "
                         "made since the last import. Almost never what you "
                         "want -- the default merges.")
    ap.add_argument("--force", action="store_true",
                    help="with --replace, allow it against a store that holds "
                         "data (DESTRUCTIVE) — only for throwaway targets")
    a = ap.parse_args()
    mode = "replace" if a.replace else "merge"
    s = run(a.workbook, a.out, force=a.force, mode=mode)
    print(f"Normalization complete ({mode}) →", a.out)
    for k, v in s.items():
        if k == "merge_report":
            continue        # skip-ok: printing the summary, not importing a row
        print(f"  {k}: {v}")
    if s.get("merge_report"):
        from merge import format_report
        print("\n" + format_report(s["merge_report"]))
