#!/usr/bin/env python3
"""Independent audit: source workbook vs the CRM store.

Deliberately does NOT reuse normalize.py's parsing — it re-reads the workbook
with the corrected semantics (confirmed by Zeeshan 2026-07-03) so it can catch
the pipeline's mistakes:

- Project Tracker holds TWO stacked tables:
    rows 2..~5   OPEN ORDERS   (A=Unrivaled Project#, B=Client PO#, C=Start,
                                D=Client, E=Location, F=Notes, G..T vendor
                                PO/ship-date pairs)
    row 12       CLIENT INVOICES header, rows 13+ data:
                 A=Invoice Number (== customer order #), B=PO#/Order Number
                 (client PO / contract), C=Invoice Date, D=payment status
                 (header says DUE Date but values are Paid/etc.), E=Client,
                 F=payment notes, G.. = EITHER pay date + "Vendors were" prose
                 OR vendor-PO shipment legs ("everything under I12 that looks
                 like a PO is a shipment number").
- Sales Tracker: A=Project#(quote#), F=Status, H=Client PO#, I=Invoice#,
  J..N financials.

Usage:
    python3 audit_workbook_vs_store.py --workbook <xlsx> --store <dir> --out <report.md>
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl

# ------------------------------------------------------------- classifiers

PO_RE = re.compile(r"(?i)(\bp\.?o\.?\s*#?\s*\d|\bpo\b|^p\d{4,})")
DATE_STR_RE = re.compile(r"^\d{4}-\d{2}-\d{2}|^\d{1,2}/\d{1,2}/\d{2,4}$")
PAID_RE = re.compile(r"(?i)\bpaid\b")


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def is_date(v):
    if isinstance(v, (datetime, date)):
        return True
    s = clean(v)
    return bool(s and DATE_STR_RE.match(s))


def is_po(v):
    s = clean(v)
    return bool(s and not is_date(v) and PO_RE.search(s))


def norm_no(v):
    s = clean(v)
    if s is None:
        return None
    return re.sub(r"\.0$", "", s)


def norm_name(s):
    s = re.sub(r"\s*\([^)]*\)\s*", " ", str(s or ""))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def close(a, b, tol=0.01):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a) == str(b)


# ------------------------------------------------------ workbook extraction

def extract(workbook):
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)

    deals = []
    for sheet in [s for s in wb.sheetnames if s.lower().startswith("sales tracker")]:
        ws = wb[sheet]
        year = 2025 if "2025" in sheet else 2026
        rows = list(ws.iter_rows(min_row=1, max_row=2200, values_only=True))
        hrow = next(i for i, r in enumerate(rows) if r and clean(r[0]) == "Project#")
        for rn, r in enumerate(rows[hrow + 1:], hrow + 2):
            if not r or (r[0] is None and (len(r) < 3 or r[2] is None)):
                continue
            deals.append({
                "sheet": sheet, "row": rn, "year": year,
                "project_no": norm_no(r[0]),
                "customer": clean(r[2]),
                "description": clean(r[3]) if len(r) > 3 else None,
                "location": clean(r[4]) if len(r) > 4 else None,
                "status": (clean(r[5]) or "").lower() or None if len(r) > 5 else None,
                "client_po_no": clean(r[7]) if len(r) > 7 else None,
                "invoice_no": norm_no(r[8]) if len(r) > 8 else None,
                "revenue": r[9] if len(r) > 9 else None,
                "total_cost": r[10] if len(r) > 10 else None,
                "gross_profit": r[11] if len(r) > 11 else None,
                "margin": r[13] if len(r) > 13 else None,
                "notes": clean(r[14]) if len(r) > 14 else None,
            })

    ws = wb["Project Tracker"]
    rows = list(ws.iter_rows(min_row=1, max_row=400, values_only=True))
    inv_header = next(i for i, r in enumerate(rows)
                      if r and clean(r[0]) == "Invoice Number:")

    open_orders = []
    for rn, r in enumerate(rows[1:inv_header], 2):
        if not r or r[0] is None:
            continue
        legs = []
        for j in range(6, 20, 2):  # G,I,K,M,O,Q,S paired with next col
            po = clean(r[j]) if len(r) > j else None
            sd = r[j + 1] if len(r) > j + 1 else None
            if po or sd:
                legs.append({"po": po, "ship_date": sd})
        open_orders.append({
            "row": rn, "key": norm_no(r[0]), "client_po": clean(r[1]),
            "start": r[2], "client": clean(r[3]), "location": clean(r[4]),
            "notes": clean(r[5]), "legs": legs,
        })

    invoices = []
    for rn, r in enumerate(rows[inv_header + 1:], inv_header + 2):
        if not r or r[0] is None:
            continue
        cells = [(j, r[j]) for j in range(6, min(len(r), 24)) if r[j] is not None]
        legs, pay_date, vendor_notes = [], None, []
        cur_po = None
        for j, v in cells:
            if is_po(v):
                if cur_po:
                    legs.append({"po": cur_po, "ship_date": None})
                cur_po = clean(v)
            elif is_date(v):
                if cur_po:
                    legs.append({"po": cur_po, "ship_date": v})
                    cur_po = None
                elif pay_date is None:
                    pay_date = v
                # else: stray date, ignore
            else:
                vendor_notes.append(clean(v))
        if cur_po:
            legs.append({"po": cur_po, "ship_date": None})
        invoices.append({
            "row": rn,
            "invoice_no": norm_no(r[0]),
            "client_po_raw": clean(r[1]) if len(r) > 1 else None,
            "invoice_date": r[2] if len(r) > 2 else None,
            "payment_status_raw": clean(r[3]) if len(r) > 3 else None,
            "client": clean(r[4]) if len(r) > 4 else None,
            "payment_notes": clean(r[5]) if len(r) > 5 else None,
            "pay_date": pay_date,
            "legs": legs,
            "vendor_notes": [v for v in vendor_notes if v],
        })

    return deals, open_orders, invoices


# ------------------------------------------------------------------- audit

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--store", required=True)
    ap.add_argument("--out", default="audit-report.md")
    ap.add_argument("--json", default="audit-findings.json")
    a = ap.parse_args()

    store = Path(a.store)
    # Explicit encodings: the store is written as UTF-8 with ensure_ascii=False,
    # and Windows' default locale codec (cp1252) can't read it back. utf-8-sig
    # also survives a BOM from any hand-edit in Notepad.
    S = {}
    for n in ["companies", "contacts", "projects", "shipments", "vendors"]:
        try:
            with open(store / f"{n}.json", encoding="utf-8-sig") as f:
                S[n] = json.load(f)
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            sys.exit(f"FATAL: {n}.json is unreadable ({e}) — fix or restore it, "
                     f"then re-run the audit")
    edited = set()
    clog = store / "changelog.jsonl"
    if clog.exists():
        with open(clog, encoding="utf-8-sig") as f:
            for line in f:
                try:
                    e = json.loads(line)
                except json.JSONDecodeError:
                    continue  # one torn line must not kill the audit
                if e.get("op") in ("update", "create"):
                    edited.add((e["entity"], str(e["key"])))

    deals, open_orders, invoices = extract(a.workbook)
    F = defaultdict(list)  # findings by category

    # ---- 1. Deal rows <-> store projects, field by field
    sp = {str(p["project_no"]): p for p in S["projects"] if p.get("project_no")}
    seen = set()
    for d in deals:
        pno = d["project_no"]
        if not pno or not re.fullmatch(r"\d{2,6}", pno):
            continue  # handled by needs_review in pipeline
        p = sp.get(pno)
        if not p:
            F["deal_missing_in_crm"].append(
                {"project_no": pno, "sheet": d["sheet"], "row": d["row"],
                 "customer": d["customer"]})
            continue
        if pno in seen:
            continue  # duplicate deal rows already flagged by pipeline
        seen.add(pno)
        was_edited = ("project", pno) in edited
        diffs = {}
        for f in ["status", "description", "location", "client_po_no",
                  "invoice_no", "notes"]:
            if (d[f] or None) != (p.get(f) or None):
                diffs[f] = {"sheet": d[f], "crm": p.get(f)}
        for f in ["revenue", "total_cost", "gross_profit", "margin"]:
            if not close(d[f], p.get(f)):
                diffs[f] = {"sheet": d[f], "crm": p.get(f)}
        if norm_name(d["customer"]) and norm_name(p.get("company_name")) \
                and norm_name(p.get("company_name")) not in norm_name(d["customer"]):
            diffs["customer"] = {"sheet": d["customer"], "crm": p.get("company_name")}
        if diffs:
            F["project_field_mismatch" if not was_edited
              else "project_differs_but_edited_in_crm"].append(
                {"project_no": pno, "sheet": d["sheet"], "row": d["row"],
                 "diffs": diffs})

    for pno, p in sp.items():
        if re.fullmatch(r"\d{2,6}", pno) and pno not in {d["project_no"] for d in deals}:
            F["crm_project_not_in_workbook"].append(
                {"project_no": pno, "company": p.get("company_name"),
                 "edited_in_crm": ("project", pno) in edited})

    # ---- 2. Invoice table -> expected collection status on projects
    inv_by_no = {}
    for inv in invoices:
        if inv["invoice_no"] and re.fullmatch(r"\d{2,6}", inv["invoice_no"]):
            inv_by_no[inv["invoice_no"]] = inv
    proj_by_inv = defaultdict(list)
    for p in S["projects"]:
        if p.get("invoice_no"):
            proj_by_inv[norm_no(p["invoice_no"])].append(p)

    joined = 0
    for ino, inv in inv_by_no.items():
        paid = bool(PAID_RE.search(inv["payment_status_raw"] or "")
                    or PAID_RE.search(inv["payment_notes"] or ""))
        expected = "paid" if paid else "open"
        targets = proj_by_inv.get(ino, [])
        if targets:
            joined += 1
            for p in targets:
                if (p.get("collection_status") or "open") != expected \
                        and p.get("collection_status") != expected:
                    F["collection_status_missing_or_wrong"].append(
                        {"invoice_no": ino, "project_no": p["project_no"],
                         "sheet_row": inv["row"], "expected": expected,
                         "crm": p.get("collection_status"),
                         "sheet_status": inv["payment_status_raw"]})
        else:
            F["invoice_unmatched_to_project"].append(
                {"invoice_no": ino, "client": inv["client"],
                 "sheet_row": inv["row"], "status": inv["payment_status_raw"]})

    # ---- 3. Shipments: bogus legs in CRM + real legs missed
    date_pos = [s for s in S["shipments"] if is_date(s.get("vendor_po_raw"))]
    prose_dates = [s for s in S["shipments"]
                   if s.get("ship_date") and not is_date(s["ship_date"])]
    for s in date_pos:
        F["bogus_shipment_po_is_a_date"].append(
            {"shipment_id": s["shipment_id"], "vendor_po_raw": s["vendor_po_raw"],
             "client": s.get("client_name")})
    for s in prose_dates:
        F["bogus_shipment_date_is_prose"].append(
            {"shipment_id": s["shipment_id"], "ship_date": str(s["ship_date"])[:60],
             "client": s.get("client_name")})

    # expected legs from corrected extraction
    expected_legs = []
    for oo in open_orders:
        for leg in oo["legs"]:
            if leg["po"]:
                expected_legs.append({"src": f"open-orders r{oo['row']}",
                                      "key": oo["key"], "po": leg["po"],
                                      "client": oo["client"]})
    for inv in invoices:
        for leg in inv["legs"]:
            expected_legs.append({"src": f"invoices r{inv['row']}",
                                  "key": inv["invoice_no"], "po": leg["po"],
                                  "client": inv["client"]})
    store_pos = {norm_name(s.get("vendor_po_raw")) for s in S["shipments"]
                 if s.get("vendor_po_raw")}
    for leg in expected_legs:
        if norm_name(leg["po"]) not in store_pos:
            F["real_shipment_leg_missing_in_crm"].append(leg)

    n_bogus = len(date_pos) + len(prose_dates)
    # store shipments whose po looks real
    real_in_store = [s for s in S["shipments"]
                     if s.get("vendor_po_raw") and not is_date(s["vendor_po_raw"])]

    # ---- 4. Contacts + vendors spot equality
    wb2 = openpyxl.load_workbook(a.workbook, read_only=True, data_only=True)
    ws = wb2["Client Contacts"]
    sheet_contacts = []
    for r in ws.iter_rows(min_row=2, max_row=1100, values_only=True):
        if r and (clean(r[0]) or clean(r[1])):
            sheet_contacts.append({"biz": clean(r[0]), "name": clean(r[1]),
                                   "email": clean(r[2]),
                                   "phone": clean(r[3]) if len(r) > 3 else None,
                                   "title": clean(r[4]) if len(r) > 4 else None})
    crm_c = {(norm_name(c.get("company_name")), norm_name(c.get("name")))
             for c in S["contacts"]}
    for c in sheet_contacts:
        if (norm_name(c["biz"]), norm_name(c["name"])) not in crm_c:
            F["contact_missing_in_crm"].append(c)

    # ---- summary + report
    summary = {
        "deal_rows_in_workbook": len([d for d in deals if d["project_no"]]),
        "projects_in_crm": len(S["projects"]),
        "invoice_rows_in_workbook": len(invoices),
        "invoices_joined_to_projects_via_invoice_no": joined,
        "open_order_rows": len(open_orders),
        "shipments_in_crm": len(S["shipments"]),
        "  of which BOGUS (pay-dates/prose parsed as legs)": n_bogus,
        "  of which look real": len(real_in_store),
        "expected_real_legs_in_workbook": len(expected_legs),
        "sheet_contacts": len(sheet_contacts),
        "contacts_in_crm": len(S["contacts"]),
        "records_edited_in_crm_since_migration": len(edited),
    }
    counts = {k: len(v) for k, v in sorted(F.items())}

    with open(a.json, "w", encoding="utf-8") as f:
        json.dump({"summary": summary, "counts": counts, "findings": F},
                  f, indent=1, default=str)

    L = ["# Workbook ↔ CRM Audit", "",
         f"_Generated {datetime.now():%Y-%m-%d %H:%M}. Semantics per Zeeshan "
         "2026-07-03: Invoice # = customer order; PO#/Order # = client PO "
         "(contract); trailing columns of the invoice table = shipment (vendor "
         "PO) numbers._", "", "## Summary", ""]
    for k, v in summary.items():
        L.append(f"- {k}: **{v}**")
    L += ["", "## Findings by category", ""]
    for k, v in sorted(F.items(), key=lambda kv: -len(kv[1])):
        L.append(f"### {k} — {len(v)}")
        L.append("")
        for item in v[:200]:
            L.append(f"- `{json.dumps(item, default=str)[:240]}`")
        if len(v) > 200:
            L.append(f"- …and {len(v) - 200} more (see JSON)")
        L.append("")
    # utf-8 explicitly: the report contains ↔/… and cp1252 can't encode them
    with open(a.out, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    print(json.dumps({"summary": summary, "counts": counts}, indent=1, default=str))


if __name__ == "__main__":
    main()
