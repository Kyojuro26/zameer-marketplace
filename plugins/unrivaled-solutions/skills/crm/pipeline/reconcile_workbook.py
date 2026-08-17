#!/usr/bin/env python3
"""Reconcile the MONEY: workbook totals vs what the CRM's own screens show.

audit_workbook_vs_store.py already compares row counts and per-field values.
It never sums anything -- so a store could agree with the workbook field by
field and still put a wrong number in the top bar, which is exactly what
happened: "close, but not right" was reported against figures nobody had ever
reconciled.

This re-reads the workbook INDEPENDENTLY (its own header lookup, its own
number coercion) rather than importing normalize.py's row loop. A checker that
shares the importer's parsing cannot catch the importer's mistakes -- the same
reason payment_words.py exists in one place and audit_workbook_vs_store.py
re-derives the sheet semantics.

What it answers, in the order the operator asks it:

  1. Does "Won revenue (YYYY)" match the sheet?         -> to the dollar, or the rows that differ
  2. Does "Pending pipeline (YYYY)" match?              -> same
  3. Does "Open receivables (YYYY)" mean anything?      -> see the caveat it prints
  4. Can the Receivables screen actually price invoices? -> the join health
  5. Is any percentage being read as a payment when it is a rep's cut?

Usage:
    python3 reconcile_workbook.py --workbook <xlsx> --store <dir> [--year 2026]
    python3 reconcile_workbook.py ... --json          machine-readable
    python3 reconcile_workbook.py ... --quiet         only problems

Exit code is 1 when a total disagrees or the invoice ledger cannot be priced,
so this can gate a release rather than just inform one.
"""
import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

TOL = 0.5          # dollars; below this is float noise, not a discrepancy


def st(v):
    return "" if v is None else str(v)


def money(v):
    return f"${v:,.0f}"


def num(v):
    """Independent of normalize.num() on purpose. Accepts what a spreadsheet
    cell actually holds: a number, "$1,200", "(500)" for negatives, "" ."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return 0.0
    try:
        f = float(s)
    except ValueError:
        return 0.0
    return -f if neg else f


# ------------------------------------------------------------- workbook ----

def read_deal_sheets(path):
    """{year: {status: total}} plus the rows behind each, read straight off the
    sheet with no help from the importer."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    try:
        for name in [s for s in wb.sheetnames if s.lower().startswith("sales tracker")]:
            year = next((int(y) for y in re.findall(r"\b(20\d\d)\b", name)), None)
            ws = wb[name]
            rows = list(ws.iter_rows(min_row=1, max_row=20000, values_only=True))
            hrow = next((i for i, r in enumerate(rows)
                         if r and st(r[0]).strip().lower().startswith("project#")), None)
            if hrow is None:
                out.setdefault(year, {"sheet": name, "error": "no Project# header row"})
                continue
            hdr = [st(c).strip().lower() for c in rows[hrow]]

            def col(*names):
                for n in names:
                    for i, h in enumerate(hdr):
                        if h.startswith(n):
                            return i
                return None

            ci = {"pno": col("project#"), "status": col("status"),
                  "rev": col("revenue"), "cust": col("customer")}
            by_status = defaultdict(float)
            rows_by_status = defaultdict(list)
            keyless = []
            seen_pno = Counter()
            for n, r in enumerate(rows[hrow + 1:], start=hrow + 2):
                if not r or not any(c is not None for c in r):
                    continue
                status = st(r[ci["status"]]).strip().lower() if ci["status"] is not None else ""
                rev = num(r[ci["rev"]]) if ci["rev"] is not None else 0.0
                pno = st(r[ci["pno"]]).strip()
                if not status and rev == 0:
                    continue          # a label/spacer row, not a deal
                by_status[status or "(blank)"] += rev
                rows_by_status[status or "(blank)"].append((n, pno, rev))
                if not pno:
                    keyless.append((n, rev))
                else:
                    seen_pno[re.sub(r"\.0$", "", pno)] += 1
            out[year] = {"sheet": name, "by_status": dict(by_status),
                         "rows": dict(rows_by_status), "keyless": keyless,
                         "dupes": {k: c for k, c in seen_pno.items() if c > 1}}
    finally:
        wb.close()
    return out


# ---------------------------------------------------------------- store ----

def read_store(d):
    d = Path(d)

    def load(n):
        p = d / f"{n}.json"
        try:
            v = json.loads(p.read_text(encoding="utf-8-sig"))
        except FileNotFoundError:
            return []
        return v if isinstance(v, list) else []
    return {n: load(n) for n in ("projects", "invoices", "companies", "shipments")}


def kpi_totals(projects, year):
    """The three figures EXACTLY as build_view.kpis() computes them, so this
    reconciles what is on the screen rather than an idealised version of it."""
    cur = [p for p in projects if st(p.get("year")) == str(year)]
    def rev(p):
        try:
            return float(p.get("revenue"))
        except (TypeError, ValueError):
            return 0.0
    won = sum(rev(p) for p in cur if st(p.get("status")) == "won")
    pend = sum(rev(p) for p in cur if st(p.get("status")) == "pending")
    recv = sum(rev(p) for p in cur
               if st(p.get("collection_status")) and st(p.get("collection_status")) != "paid")
    return {"won": won, "pending": pend, "receivables": recv, "n": len(cur),
            "with_collection_status": sum(1 for p in cur if st(p.get("collection_status")))}


def invoice_health(inv, projects):
    """Can the Receivables screen put a number against an invoice at all?

    Mirrors build_view.invoiceAmount(): an invoice with no project_no has NO
    amount (not zero), and the project must belong to the same company. If most
    invoices cannot be priced, the screen's total is not a receivables figure
    however correct its arithmetic is."""
    idx = {}
    for p in projects:
        if st(p.get("project_no")):
            idx.setdefault((st(p.get("company_id")), st(p.get("project_no"))), p)

    def amount(i):
        pno = st(i.get("project_no"))
        if not pno:
            return None
        p = idx.get((st(i.get("company_id")), pno))
        if not p:
            return None
        try:
            return float(p.get("revenue"))
        except (TypeError, ValueError):
            return None
    unpaid = [i for i in inv if not st(i.get("payment_status")).lower().startswith("paid")]
    priced = [i for i in inv if amount(i) is not None]
    unpaid_priced = [i for i in unpaid if amount(i) is not None]
    return {
        "invoices": len(inv), "unpaid": len(unpaid),
        "priced": len(priced), "unpriced": len(inv) - len(priced),
        "unpaid_priced": len(unpaid_priced),
        "unpaid_unpriced": len(unpaid) - len(unpaid_priced),
        "screen_total": sum(amount(i) for i in unpaid_priced),
        "with_project_no": sum(1 for i in inv if st(i.get("project_no"))),
    }


def suspicious_percentages(inv):
    """A "NN%" that is really a rep's cut, stored as a part payment.

    Deliberately its OWN pattern, not normalize.commission_like: a checker that
    imports the importer's guard agrees with it by construction and can never
    report that the guard missed a spelling. This one is broad on purpose --
    every hit is a question for a person, not a verdict."""
    pat = re.compile(r"(?:\b[A-Z]{1,4}\s*@\s*\d{1,3}\s*%"
                     r"|%\s*(?:to|for)\s+[A-Za-z]"
                     r"|\b(?:comm|commission|split|payout|rep)\b[^.]{0,20}\d{1,3}\s*%"
                     r"|\d{1,3}\s*%[^.]{0,20}\b(?:comm|commission|split|payout|rep)\b)",
                     re.IGNORECASE)
    hits = []
    for i in inv:
        if not st(i.get("payment_status")).startswith("partial"):
            continue
        blob = f"{st(i.get('payment_status_raw'))} {st(i.get('payment_notes'))}"
        if pat.search(blob):
            hits.append({"invoice_no": st(i.get("invoice_no")),
                         "stored_as": st(i.get("payment_status")),
                         "text": blob.strip()[:90]})
    return hits


# --------------------------------------------------------------- report ----

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--store", required=True)
    ap.add_argument("--year", type=int, default=None,
                    help="KPI year (default: the latest Sales Tracker sheet)")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--quiet", action="store_true", help="print only problems")
    a = ap.parse_args()

    sheets = read_deal_sheets(a.workbook)
    store = read_store(a.store)
    years = sorted(y for y in sheets if y)
    year = a.year or (years[-1] if years else None)
    if year is None:
        sys.exit("no Sales Tracker year sheet found in the workbook")

    wb_year = sheets.get(year, {})
    wb_status = wb_year.get("by_status", {})
    crm = kpi_totals(store["projects"], year)
    health = invoice_health(store["invoices"], store["projects"])
    sus = suspicious_percentages(store["invoices"])

    problems = []
    lines = []
    P = lines.append

    P(f"Reconciling {Path(a.workbook).name}  ->  {a.store}")
    P(f"KPI year: {year}   (sheet: {wb_year.get('sheet', '?')})")
    P("")
    P("TOP-BAR FIGURES")
    for label, key in (("Won revenue", "won"), ("Pending pipeline", "pending")):
        w, c = wb_status.get(key, 0.0), crm[key]
        ok = abs(w - c) < TOL
        P(f"  {label:<20} workbook {money(w):>14}   CRM {money(c):>14}   "
          f"{'match' if ok else 'DIFFERS by ' + money(c - w)}")
        if not ok:
            problems.append(f"{label} differs by {money(c - w)}")
            for st_name, rows in sorted(wb_year.get("rows", {}).items()):
                if st_name != key:
                    continue
                P(f"      workbook rows under '{key}': {len(rows)}")
    P("")
    P("  Open receivables      "
      f"CRM {money(crm['receivables']):>14}   -- NOT reconcilable against the sheet")
    P(f"      it sums {year} PROJECTS whose collection_status is set and not 'paid';")
    P(f"      only {crm['with_collection_status']} of {crm['n']} such projects have one at all,")
    P("      so the figure is those few rows, not what is owed.")
    if crm["n"] and crm["with_collection_status"] / max(crm["n"], 1) < 0.5:
        problems.append(
            f"Open receivables is derived from collection_status, present on only "
            f"{crm['with_collection_status']}/{crm['n']} {year} projects")

    P("")
    P("RECEIVABLES SCREEN (unpaid invoices priced from their linked project)")
    P(f"  invoices {health['invoices']}   unpaid {health['unpaid']}   "
      f"carrying a project_no {health['with_project_no']}")
    P(f"  can be priced {health['priced']}   CANNOT {health['unpriced']}")
    P(f"  screen total {money(health['screen_total'])} from {health['unpaid_priced']} "
      f"priced unpaid invoice(s)")
    if health["invoices"] and health["unpriced"] / health["invoices"] > 0.25:
        problems.append(
            f"{health['unpriced']} of {health['invoices']} invoices cannot be priced "
            f"(no project link), so the Receivables screen is mostly empty")

    P("")
    P("PERCENTAGES THAT MAY BE A REP'S CUT, STORED AS A PART PAYMENT")
    if not sus:
        P("  none")
    else:
        problems.append(f"{len(sus)} invoice(s) stored as part-paid from a "
                        f"commission-shaped percentage")
        for h in sus[:15]:
            P(f"  inv {h['invoice_no']:<10} {h['stored_as']:<14} {h['text']!r}")
        if len(sus) > 15:
            P(f"  ... and {len(sus) - 15} more")

    other = {k: v for k, v in wb_status.items() if k not in ("won", "pending")}
    if other and not a.quiet:
        P("")
        P(f"OTHER STATUSES IN THE {year} SHEET (not shown in the top bar)")
        for k in sorted(other, key=lambda k: -other[k]):
            P(f"  {k:<20} {money(other[k]):>14}")
    if wb_year.get("keyless") and not a.quiet:
        tot = sum(r for _, r in wb_year["keyless"])
        P("")
        P(f"WORKBOOK ROWS WITH NO PROJECT NUMBER: {len(wb_year['keyless'])} "
          f"carrying {money(tot)}")
        P("  these still total into the figures above; they just cannot be opened "
          "in the CRM by number")
    if wb_year.get("dupes"):
        P("")
        P(f"DUPLICATE PROJECT NUMBERS IN THE {year} SHEET: "
          f"{', '.join(sorted(wb_year['dupes'])[:10])}")
        problems.append(f"{len(wb_year['dupes'])} duplicate project number(s) in the sheet")
    for y in years:
        if y == year:
            continue
        s = sheets[y].get("by_status", {})
        if not a.quiet:
            P("")
            P(f"FOR REFERENCE -- {y} sheet, excluded from the {year} top bar: "
              f"won {money(s.get('won', 0))}, pending {money(s.get('pending', 0))}")

    P("")
    if problems:
        P(f"{len(problems)} PROBLEM(S):")
        for p in problems:
            P(f"  - {p}")
    else:
        P("No discrepancies found.")

    if a.json:
        print(json.dumps({"year": year, "workbook": wb_status, "crm": crm,
                          "invoice_health": health, "suspicious": sus,
                          "problems": problems}, indent=2, default=str))
    else:
        print("\n".join(lines))
    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
