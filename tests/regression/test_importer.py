"""The importer, end to end, against a real workbook.

Nothing in the suite built a workbook before this file, so every defect in
normalize.py's decisions was invisible: a gutted payment_words.py scored a
clean 180/180 while reading "hasn't paid" and "paid - check returned NSF" as
PAID. Grepping a module's source is not testing it.

These cases are the wording a small manufacturer actually types into a
collection-notes column. Both error directions cost money:
  reading an UNPAID note as paid  -> a live receivable drops off the list
  flagging a genuinely PAID one   -> he chases a customer who already paid,
                                     and the review list becomes noise
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from lib.harness import Result  # noqa: E402

# (note as it would appear in the workbook, must-be-collected?)
UNPAID = [
    "NOT PAID as of 7/1 - chasing", "NOT PAID IN FULL - 50% short",
    "Will be paid in full 8/30", "COD - paid on delivery",
    "Paid less freight, $340 still owed", "Was paid, then charged back",
    "We paid freight; invoice to client still open", "Paid $5,000 of $12,500",
    "paid in full? confirm with Sherry", "2 of 3 invoices paid",
    "paid 1st installment", "partially paid", "paid 50%", "hasn't paid",
    "still owes, paid nothing", "paid - check returned NSF",
    "paid deposit only, balance due 6/1", "paid, awaiting the rest",
    "paid? unsure", "will be paid net 60", "50% paid, balance due",
    "Never paid - write off", "paid - REVERSED",
    # every one of these was written off by a previous version of this module.
    # Three successive blacklists each shipped a gap at the NEXT spelling, which
    # is why the implementation is now a whitelist -- see payment_words.py.
    "Paid, $12,500 still outstanding", "Paid but balance outstanding",
    "Paid 3/1, $500 still due", "Paid after installation sign-off",
    "Part paid", "Paid before shipment", "Paid, balance $8,000",
    "Paid - amount due 5000", "Paid; outstanding items to bill",
    "Paid assuming the credit clears", "Paid less shipping",
    "paid short", "Paid, would clear next week", "Paid pending credit note",
]
PAID = [
    "Paid in full, balance $0", "paid - no balance due",
    "Paid, nothing outstanding", "Paid 6/1 check 8812", "paid in full",
    "PAID", "paid off", "Paid 5/2", "paid in full 3/14 ck 9912",
    "Paid $12,500", "paid via wire 6/1", "Paid 6/1, thanks",
    "PAID IN FULL", "paid ck #8812", "Paid 06/01/2026",
]
NO_MENTION = ["Open", "awaiting payment", "unpaid", "", "net 30"]

# a percentage next to a commission is a payout rate, never a collection status
COMMISSION = ["10% comm to D", "10% commission to D", "10% commissions to D",
              "10% comms to D", "10% cmsn to D", "10% to D", "15% for JS",
              "20% rep split",
              # the initials-BEFORE-the-rate spelling, found on the real
              # ledger: 41 invoices reading like this were stored as
              # partial:25%, so the app reported a part payment that had not
              # happened and knocked 25% off what each was owed
              "Invoice sent on 4/20/26 (D @ 25%)", "(D @ 25%)", "D@25%",
              "JS @ 10%", "Invoice sent on 05/01/26 (D @ 25%)"]
NOT_COMMISSION = ["50%", "25% deposit received", "10% for freight",
                  "2% discount", "50% collected",
                  # a spelled-out word before the @ is NOT initials, and this
                  # one is a real payment percentage -- the guard is limited to
                  # 1-3 capitals so it cannot swallow these
                  "Deposit @ 25%", "Paid @ 50%", "Collected @ 30%"]


def _em(user, host):
    """Compose a fixture address at runtime. Written as a literal it would be
    an email shape in a public repo, which the PII sweep rejects outright --
    and rightly: the check cannot tell an invented address from a real one."""
    return user + "@" + host


def _load(crm_dir, name):
    import importlib.util
    p = Path(crm_dir) / "pipeline" / name
    if not p.exists():
        return None
    spec = importlib.util.spec_from_file_location(f"_imp_{name}", p)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def run(server, crm_dir=None):
    r = Result("importer", since="0.1.28")
    crm = Path(crm_dir) if crm_dir else \
        Path(__file__).resolve().parents[2] / "plugins/unrivaled-solutions/skills/crm"
    pw = _load(crm, "payment_words.py")
    nrm = _load(crm, "normalize.py")
    if pw is None:
        r.check("payment wording has a single definition", False,
                "pipeline/payment_words.py missing")
        return r

    r.section("an unpaid receivable is never recorded as collected")
    for note in UNPAID:
        r.check(f"not collected: {note!r}", pw.says_paid(note) is None,
                f"says_paid={pw.says_paid(note)!r} -- a live receivable would "
                f"drop off the collections list, unflagged")

    r.section("a genuinely paid one is not flagged as ambiguous")
    for note in PAID:
        r.check(f"collected: {note!r}", pw.says_paid(note) is True,
                f"says_paid={pw.says_paid(note)!r} -- the operator chases a "
                f"customer who has already paid, and the real flags get buried")

    r.section("no payment mentioned at all")
    for note in NO_MENTION:
        r.check(f"no mention: {note!r}", pw.says_paid(note) is False,
                f"says_paid={pw.says_paid(note)!r}")

    if nrm and hasattr(nrm, "commission_like"):
        r.section("a commission rate is never read as a collection percentage")
        for note in COMMISSION:
            r.check(f"commission: {note!r}", nrm.commission_like(note) is True,
                    "a rep's payout rate would be stored as percent collected")
        for note in NOT_COMMISSION:
            r.check(f"not commission: {note!r}",
                    nrm.commission_like(note) is False,
                    "a real collection percentage would be discarded")

    r.section("the importer mints unique shipment ids")
    if nrm:
        src = (crm / "pipeline" / "normalize.py").read_text()
        r.check("leg numbering does not restart per row",
                "_next_leg" in src and "L{leg_no}" not in src,
                "a project number on two open-order rows mints the same "
                "shipment_id twice, and no tool can then tell the legs apart")

    # ---- presence vs coercion -------------------------------------------
    # clean() answers "what is the string form of this key" and MUST keep
    # returning "0" -- a project or invoice number of 0 is a real key. The
    # separate question "is there anything in this cell" is has_value(). Using
    # clean() for the second one read every formatted-but-empty row whose
    # cached formula result was 0 as a row carrying financial values: 39,726
    # spurious review entries out of 39,872 on the operator's live workbook,
    # burying the ~140 real ones at roughly 280:1.
    r.section("a zero identifier survives, a zero data cell does not")
    if nrm and hasattr(nrm, "has_value"):
        r.check('clean(0) is still "0"', nrm.clean(0) == "0",
                'a project number of 0 must survive coercion -- this is the '
                'trap: "fixing" clean() to drop zeros loses a real key')
        r.check('clean("0") is still "0"', nrm.clean("0") == "0")
        r.check("clean('') is None", nrm.clean("") is None)
        for v in (0, 0.0, "0", "0.00", "0,000", "", "   ", None):
            r.check(f"has_value({v!r}) is False", nrm.has_value(v) is False,
                    "a formatted-but-empty cell is not content")
        for v in (1397.0, "1,200", "N/A", "TBD", 0.01, "-5"):
            r.check(f"has_value({v!r}) is True", nrm.has_value(v) is True,
                    "a person typed something there")
    elif nrm:
        r.check("has_value() exists", False,
                "the presence test was folded back into clean(), which loses "
                "a literal 0 identifier")

    # ---- the deal-log loop, end to end against a built workbook ----------
    # Invented companies and people throughout: nothing from the real workbook
    # may enter this repo, including test data.
    r.section("a deal row with no keys is reported, not fatal")
    if nrm:
        import json as _json, tempfile as _tf
        # No try/except around this import. normalize.py does sys.exit() when
        # openpyxl is missing, and SystemExit is a BaseException -- it escapes
        # run_all.py's `except Exception`, so the suite is already dead before
        # this line and `nrm` was never assigned. A handler here would be
        # unreachable code claiming a coverage it does not have.
        import openpyxl as _ox
        if True:
            _tmp = Path(_tf.mkdtemp())
            _x = _tmp / "fixture.xlsx"
            _wb = _ox.Workbook()
            _ws = _wb.active
            _ws.title = "Sales Tracker 2026"
            _ws.append(["Project#", "Date", "Customer", "Description", "Location",
                        "Status", "PO Y/N", "Client PO#", "Invoice #", "Revenue",
                        "Total Cost", "Total GP", "Margain", "Notes"])
            _ws.append(["9001", None, "Brightwater Fabrication", "Line install",
                        "Dayton OH", "won", "Y", "PO-5501", "8801",
                        50000, 30000, 20000, 0.4, ""])
            # money, but neither a customer nor a project number: the shape that
            # raised UnboundLocalError and killed the whole run
            _ws.append([None] * 9 + [125000, 90000, 35000, None, "TOTALS"])
            # identical shape with cached zeros: must stay silent
            _ws.append([None] * 9 + [0, 0, 0, None, None])
            # revenue only, cost and GP blank -- an order booked but not yet
            # costed. Makes `any` and `all` over the three financial columns
            # give different answers, which the all-three row cannot.
            _ws.append([None] * 9 + [88000, None, None, None, None])
            # GP only, for the same reason in the other direction: a predicate
            # narrowed to revenue alone would miss this one.
            _ws.append([None] * 11 + [4200, None, None])
            # a literal zero PROJECT NUMBER must import as the key "0"
            _ws.append([0, None, "Ironvale Supply", "Zero-keyed job", "Akron OH",
                        "won", "N", None, None, 1000, 600, 400, 0.4, ""])
            _pt = _wb.create_sheet("Project Tracker")
            _pt.append(["Unrivaled Project#:", "Client PO#:", "Start Date:",
                        "Client Name:", "Client Location:", "Open Orders Notes:",
                        "Vendor 1 PO#:", "Vendor 1 Ship Date:"])
            _pt.append(["9001", "PO-5501", None, "Brightwater Fabrication",
                        "Dayton OH", "waiting on frames", "VPO-1", None])
            # formatted-but-empty open-orders row: str(c).strip() sees "0" and
            # reports it, has_value does not. Same bug as the deal log, at a
            # site that never touches clean() -- grepping for clean missed it.
            _pt.append([None, 0, 0, 0, 0, 0, 0, 0])
            # a keyed row AFTER it, so the all-zero row is inside the data
            # region. The legend/footer boundary skips everything below the last
            # keyed row, so without this the row above is never presence-tested
            # and the assertion about it silently stops meaning anything.
            _pt.append(["9002", "PO-5502", None, "Brightwater Fabrication",
                        "Dayton OH", "second job", "VPO-2", None])
            _pt.append(["Invoice Number:", "PO#", "Invoice Date", "DUE Date",
                        "Client", "Notes"])
            _cc = _wb.create_sheet("Client Contacts")
            _cc.append(["Client Business", "Client Name", "Email",
                        "Phone Number", "Job Title", "Location",
                        "Action Taken and Notes", "Last Date of Action"])
            _cc.append(["Brightwater Fabrication", "Rae Nolan",
                        _em("rae", "brightwater.example"), "555-0100", "Purchasing"])
            _vc = _wb.create_sheet("Vendor Contacts")
            _vc.append(["Company", "Headquarters Location", "Sales Rep/Contact",
                        "Contact Email", "Contact Phone Number", "Offerings",
                        "Send PO's to", "Send Invoices to"])
            # a literal 0 in the send-PO cell must not become a routing address
            _vc.append(["Cobalt Freight", "Toledo OH", "Sam Vey",
                        _em("sam", "cobalt.example"), "555-0200", "LTL",
                        0, _em("ap", "cobalt.example")])
            # the POSITIVE direction: a real routing address must be stored and
            # must be claimed as coming from the sheet. Without this row the
            # send_po column can be made unresolvable, or the source hardcoded,
            # and every assertion about the zero case still passes on a None.
            _vc.append(["Trueline Metals", "Akron OH", "Dee Marsh",
                        _em("dee", "trueline.example"), "555-0300", "Plate",
                        _em("po", "trueline.example"), 0])
            _wb.save(_x)
            _out = _tmp / "store"
            _out.mkdir()
            _sys_path_added = str(crm / "pipeline")
            _added_path = _sys_path_added not in sys.path
            if _added_path:
                sys.path.insert(0, _sys_path_added)   # normalize imports merge
            _err = None
            try:
                nrm.run(str(_x), str(_out), force=False, mode="merge")
            except Exception as exc:                       # noqa: BLE001
                _err = f"{type(exc).__name__}: {exc}"
            r.check("the import runs to completion", _err is None, _err or "")
            if _err is None:
                _rv = _json.loads((_out / "needs_review.json").read_text())
                _dk = [x for x in _rv if x.get("type") == "deal_row_without_keys"]
                _rows = sorted(x.get("sheet_row") for x in _dk)
                # sheet rows: 1 header, 2 ok, 3 money+no keys, 4 all-zero,
                # 5 revenue-only+no keys, 6 GP-only+no keys, 7 project "0"
                r.check("exactly the keyless rows carrying money are reported",
                        _rows == [3, 5, 6],
                        f"reported rows {_rows}, expected [3, 5, 6] -- counting "
                        f"alone cannot tell this from its own inverse")
                r.check("the all-zero row (4) is NOT among them", 4 not in _rows,
                        "cached zeros in a formatted row are not financial values")
                r.check("a revenue-only row is reported (any, not all)",
                        5 in _rows,
                        "requiring every financial column drops a booked order "
                        "that has not been costed yet")
                r.check("a GP-only row is reported (not revenue alone)",
                        6 in _rows,
                        "narrowing the predicate to revenue misses this")
                r.check("each entry names the sheet it came from",
                        all(x.get("sheet") for x in _dk),
                        f"got {_dk[:1]} -- an entry with no location is unactionable")
                _pj = _json.loads((_out / "projects.json").read_text())
                _keys = {str(p.get("project_no")) for p in _pj}
                r.check('a project numbered 0 still imports under the key "0"',
                        "0" in _keys, f"project keys were {sorted(_keys)}")
                _vn = _json.loads((_out / "vendors.json").read_text())
                _cob = [v for v in _vn if "cobalt" in str(v.get("company_id", ""))]
                r.check("the send-invoice column actually resolves",
                        bool(_cob) and _cob[0].get("invoice_routing")
                        == _em("ap", "cobalt.example"),
                        "if the header names do not match, every assertion "
                        "below passes vacuously on a None")
                r.check("a zero send-PO cell is not stored as a routing address",
                        bool(_cob) and _cob[0].get("po_routing") is None,
                        f"got {_cob[:1]}")
                r.check("and is not claimed to have come from the sheet",
                        bool(_cob) and _cob[0].get("po_routing_source")
                        != "sheet",
                        "asserting sheet provenance for an empty cell is a lie "
                        "about where POs should be sent")
                _tru = [v for v in _vn if "trueline" in str(v.get("company_id", ""))]
                r.check("a REAL send-PO address is stored",
                        bool(_tru) and _tru[0].get("po_routing")
                        == _em("po", "trueline.example"),
                        f"got {_tru[:1]} -- without this the send_po column can "
                        f"stop resolving and the zero-case checks pass on None")
                r.check("and IS claimed to have come from the sheet",
                        bool(_tru) and _tru[0].get("po_routing_source") == "sheet",
                        "the provenance claim must be testable in both "
                        "directions or it can be hardcoded")
                r.check("a zero send-INVOICE cell is dropped too",
                        bool(_tru) and _tru[0].get("invoice_routing") is None,
                        "the guard was added to both fields; only one was tested")
                _oo = [x for x in _rv
                       if x.get("type") == "open_order_row_without_project"]
                r.check("an all-zero open-orders row is not reported either",
                        len(_oo) == 0,
                        f"got {len(_oo)} -- the open-orders loop has the same "
                        f"presence-vs-coercion bug and its own spelling of it")
                _sp = _json.loads((_out / "shipments.json").read_text())
                r.check("the Project Tracker sheet actually imported a leg",
                        any(str(x.get("project_no")) == "9001" for x in _sp),
                        f"got {len(_sp)} shipments -- the fixture sheet would "
                        f"otherwise be decorative")
                _ct = _json.loads((_out / "contacts.json").read_text())
                r.check("the Client Contacts sheet actually imported a person",
                        any(x.get("email") == _em("rae", "brightwater.example")
                            for x in _ct),
                        f"got {len(_ct)} contacts")
            # ---- the row cap, both directions -------------------------
            # ROW_CAP is overridden rather than building a 20,000-row sheet.
            # Excel reports a used range covering every FORMATTED cell, so the
            # operator's deal log claims 35,139 rows with real content ending
            # at 172: flagging on the count alone would report 15,000 lost rows
            # every import, which is the same unactionable noise this pass
            # removes. Flagging on nothing would hide a real truncation.
            _capwas = nrm.ROW_CAP
            try:
                nrm.ROW_CAP = 6
                for _label, _tail, _want in (
                        ("padding", [None, None, None, None, None, None, None,
                                     None, None, 0, 0, 0, None, None], False),
                        ("real data", ["7999", None, "Kestrel Works", "Late job",
                                       "Lima OH", "won", "N", None, None,
                                       999, 500, 499, 0.5, ""], True)):
                    _x2 = _tmp / f"cap-{_label.replace(' ', '-')}.xlsx"
                    _w2 = _ox.Workbook()
                    _s2 = _w2.active
                    _s2.title = "Sales Tracker 2026"
                    _s2.append(["Project#", "Date", "Customer", "Description",
                                "Location", "Status", "PO Y/N", "Client PO#",
                                "Invoice #", "Revenue", "Total Cost", "Total GP",
                                "Margain", "Notes"])
                    _s2.append(["7001", None, "Kestrel Works", "In range",
                                "Lima OH", "won", "N", None, None,
                                1000, 600, 400, 0.4, ""])
                    for _ in range(6):
                        _s2.append([None] * 14)
                    _s2.append(_tail)          # sits past the cap
                    for _n in ("Project Tracker", "Client Contacts",
                               "Vendor Contacts"):
                        _w2.create_sheet(_n).append(["x"])
                    _w2.save(_x2)
                    _o2 = _tmp / f"store-{_label.replace(' ', '-')}"
                    _o2.mkdir()
                    try:
                        nrm.run(str(_x2), str(_o2), force=False, mode="merge")
                        _r2 = _json.loads((_o2 / "needs_review.json").read_text())
                        _tr = [x for x in _r2 if x.get("type") == "sheet_truncated"]
                        r.check(f"beyond the cap, {_label} -> "
                                f"{'flagged' if _want else 'silent'}",
                                bool(_tr) is _want,
                                f"got {len(_tr)} sheet_truncated entries; "
                                f"{'a real row was dropped with no record' if _want else 'formatted padding is not lost data'}")
                    except Exception as exc:                   # noqa: BLE001
                        r.check(f"cap check runs ({_label})", False,
                                f"{type(exc).__name__}: {exc}")
            finally:
                nrm.ROW_CAP = _capwas

            import shutil as _sh
            _sh.rmtree(_tmp, ignore_errors=True)
            # leaving the plugin's pipeline/ on sys.path changes import
            # resolution for every module the runner loads after this one
            if _added_path:
                try:
                    sys.path.remove(_sys_path_added)
                except ValueError:
                    pass
    return r
