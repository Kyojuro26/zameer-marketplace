"""The Live Tracker's IMPORT half: status decoded from a cell FILL COLOUR.

Nothing else in this system reads formatting. Every other field is a value in a
cell, and `values_only=True` -- which the importer used everywhere -- never
materialises a cell object, so the fill was invisible to five releases. That
makes this module's failure mode unlike the rest of the suite's: the import
still succeeds, every count still looks right, and the single most meaningful
field on the sheet is simply absent or, worse, wrong.

Four decisions are asserted here because each one has an obvious alternative
that is silently wrong:

 1. THE LEGEND IS FOUND BY STRUCTURE, NOT BY COLOUR. The legend rows sit below
    the last row carrying a project key. Scanning for "rows with a bucket
    colour" instead reads the first coloured *data* row as the legend -- and
    then reads the actual legend rows as live projects. The fixture below has
    exactly that trap in it (row 7): a real, unkeyed, cyan data row above the
    boundary. A colour-anchored implementation passes every count and gets
    both halves backwards.

 2. FF00FF00 IS REGISTERED AS NEVER-A-BUCKET. It marks the second table's
    section headers. With no entry it is an unknown colour and raises a review
    flag on every single import, which is how a review list becomes noise.

 3. AN UNKNOWN COLOUR IS NEVER GUESSED. Nearest-colour matching is the obvious
    thing to do and it would file a job under the wrong person's name with no
    trace. Status stays unset and the entry names the ARGB and the row.

 4. THE NOTE BELONGS TO THE PROJECT. It used to live only on shipments, copied
    onto every leg -- five copies on the busiest row, so editing one left four
    disagreeing -- and no copy at all on a row with no legs. Row 3 of the
    fixture is that row: a live project, a real note, zero vendor legs.

Fixture names are invented. This repo is PUBLIC and the whole tree is swept.
The bucket LABELS in particular are read from the sheet at import and stored,
never written into source -- on the real workbook they name people.
"""
import json
import re
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from lib.harness import Result  # noqa: E402

MAGENTA = "FFFF00FF"        # -> action_admin
YELLOW = "FFFFFF00"         # -> action_owner
CYAN = "FF00FFFF"           # -> awaiting_materials
GREEN = "FF00FF00"          # registered as explicitly NOT a bucket
UNKNOWN = "FFAB12CD"        # in no table at all -- must never be guessed

# Invented legend text. On the real sheet these name people.
L_ADMIN = "Waiting on the office"
L_OWNER = "With the rep"
L_AWAIT = "Waiting on materials"

NOTE_NO_LEGS = "Quote re-issued 7/2; customer still deciding on the guarding."
NOTE_FIVE_LEGS = ("Frames from two suppliers, one short-shipped. Chase the "
                  "balance before the crate ships.")


def _load(crm_dir, name):
    import importlib.util
    p = Path(crm_dir) / "pipeline" / name
    if not p.exists():
        return None
    spec = importlib.util.spec_from_file_location(f"_lt_{name}", p)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def _tracker_row(key, client, note, legs=(), client_po=None):
    """One Open Orders row. A=key, B=client PO, C=start, D=client, E=location,
    F=note, then G.. as vendor PO / ship-date pairs.

    The client PO always carries digits, as a real one does: pick_client scans
    B through E for the first name-shaped cell, and a digitless "PO-X" reads as
    a company name rather than as a PO."""
    row = [key, client_po or f"PO-{key or '9090'}", None, client,
           "Dayton OH", note]
    for po, date in legs:
        row += [po, date]
    return row


def _build_workbook(path, legend=(L_ADMIN, L_OWNER, L_AWAIT)):
    """The Project Tracker sheet, with fills, laid out exactly like the real
    one: data rows, then a blank, then the legend, then the second table.

    `legend` is the three labels in bucket order. A None entry writes the
    coloured legend row with no text in it -- a bucket the sheet never names."""
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Tracker 2026"
    ws.append(["Project#", "Date", "Customer", "Description", "Location",
               "Status", "PO Y/N", "Client PO#", "Invoice #", "Revenue",
               "Total Cost", "Total GP", "Margain", "Notes"])
    for pno, cust, desc in (
            ("5001", "Brightwater Fabrication", "Conveyor upgrade"),
            ("5002", "Brightwater Fabrication", "Guarding package"),
            ("5003", "Ironvale Supply", "Frame set"),
            ("5004", "Ironvale Supply", "Spare parts"),
            ("5005", "Kestrel Works", "Retrofit"),
            ("5006", "Kestrel Works", "Empty-note job")):
        ws.append([pno, None, cust, desc, "Dayton OH", "won", "N", None, None,
                   10000, 6000, 4000, 0.4, ""])

    pt = wb.create_sheet("Project Tracker")
    pt.append(["Unrivaled Project#:", "Client PO#:", "Start Date:",
               "Client Name:", "Client Location:", "Open Orders Notes:",
               "Vendor 1 PO#:", "Vendor 1 Ship Date:"])
    # row 2 -- magenta, matches project 5001
    pt.append(_tracker_row("5001", "Brightwater Fabrication", "On the bench",
                           [("VPO-1", "2026-09-01")]))
    # row 3 -- yellow, a real note and NO vendor legs at all. This is the row
    # the old shipment-only note could not represent.
    pt.append(_tracker_row("5002", "Brightwater Fabrication", NOTE_NO_LEGS))
    # row 4 -- cyan, five legs, the fifth in column U/V (vendor 8). The old
    # loop stopped at S/T and dropped a filled U in silence.
    pt.append(_tracker_row("5003", "Ironvale Supply", NOTE_FIVE_LEGS, [
        ("VPO-A", "2026-09-02"), ("VPO-B", "2026-09-03"),
        ("VPO-C", "2026-09-04"), ("VPO-D", "2026-09-05"),
        (None, None), (None, None), (None, None), ("VPO-U8", "2026-09-06"),
    ]))
    # row 5 -- a colour in no table. Must not be guessed at.
    pt.append(_tracker_row("5004", "Ironvale Supply", "Unknown fill here"))
    # row 6 -- GREEN. Registered as never-a-bucket, so: no status, and no flag.
    pt.append(_tracker_row("5005", "Kestrel Works", "Green section marker"))
    # row 7 -- yellow, and its notes cell is EMPTY. Status is the fill, not the
    # text, so it must still bucket -- and no note must be written over it.
    pt.append(_tracker_row("5006", "Kestrel Works", None))
    # row 8 -- THE TRAP. A real data row, cyan, with content but no project
    # number, sitting ABOVE the boundary. A colour-anchored legend scan reads
    # THIS as the legend and the rows below as live work.
    pt.append(_tracker_row(None, "Meridian Corp", "No number on this one yet",
                           [("VPO-M1", "2026-09-08")]))
    # row 9 -- keyed, but the deal log has nothing answering to it. LAST KEYED
    # ROW, so this is where the boundary sits.
    pt.append(_tracker_row("5999", "Ironvale Supply", "Keyed but unmatched",
                           [("VPO-Z", None)]))
    pt.append([None] * 8)                                       # row 10, blank
    for lab in legend:                                          # rows 11-13
        pt.append([None] * 5 + [lab])
    pt.append([None] * 5 + ["CLIENT INVOICES BELOW"])           # row 14 header
    pt.append(["Invoice Number:", "PO#", "Invoice Date", "DUE Date",
               "Client", "Notes"])                              # row 15

    for row, argb in ((2, MAGENTA), (3, YELLOW), (4, CYAN), (5, UNKNOWN),
                      (6, GREEN), (7, YELLOW), (8, CYAN), (9, CYAN),
                      (11, MAGENTA), (12, YELLOW), (13, CYAN), (14, GREEN)):
        pt.cell(row=row, column=6).fill = PatternFill(
            start_color=argb, end_color=argb, fill_type="solid")

    cc = wb.create_sheet("Client Contacts")
    cc.append(["Client Business", "Client Name", "Email", "Phone Number",
               "Job Title", "Location", "Action Taken and Notes",
               "Last Date of Action"])
    vc = wb.create_sheet("Vendor Contacts")
    vc.append(["Company", "Headquarters Location", "Sales Rep/Contact",
               "Contact Email", "Contact Phone Number", "Offerings",
               "Send PO's to", "Send Invoices to"])
    wb.save(path)


def _build_edge_workbook(path, *, trailing_unkeyed=False, stray_col_a=False,
                         theme_row=False, default_fill_row=False,
                         keyed_bodyless_unknown=False, footer_stray_leg=False,
                         legend_unreadable=False):
    """A minimal sheet for the boundary and fill-decode edge cases.

    Kept separate from the main fixture: each flag here changes what the
    boundary or the decoder is being asked, and folding them into one sheet
    would make a failure ambiguous about which rule broke.
    """
    import openpyxl
    from openpyxl.styles import Color, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Tracker 2026"
    ws.append(["Project#", "Date", "Customer", "Description", "Location",
               "Status", "PO Y/N", "Client PO#", "Invoice #", "Revenue",
               "Total Cost", "Total GP", "Margain", "Notes"])
    for pno in ("6001", "6002", "6003"):
        ws.append([pno, None, "Brightwater Fabrication", f"Job {pno}",
                   "Dayton OH", "won", "N", None, None,
                   10000, 6000, 4000, 0.4, ""])

    pt = wb.create_sheet("Project Tracker")
    pt.append(["Unrivaled Project#:", "Client PO#:", "Start Date:",
               "Client Name:", "Client Location:", "Open Orders Notes:",
               "Vendor 1 PO#:", "Vendor 1 Ship Date:"])
    pt.append(_tracker_row("6001", "Brightwater Fabrication", "on the bench"))
    pt.append(_tracker_row("6002", "Brightwater Fabrication", "theme fill here"))
    # row 4 is the LAST row that is both keyed and carries a body, so the
    # boundary sits here and everything below is footer as far as position goes
    pt.append(_tracker_row("6003", "Brightwater Fabrication", "default fill"))
    if keyed_bodyless_unknown:
        # Keyed but with nothing else on it, so it does not move the boundary
        # -- and therefore sits below it. Its fill is still unrecognised and
        # its flag must still fire: a row with a key is a data row wherever it
        # sits, and tying the flag to the boundary lost both the status and
        # any mention of it.
        pt.append(["6004"] + [None] * 7)
    elif trailing_unkeyed:
        # A real job appended at the bottom before it has a number -- the
        # normal way a row gets added. It sits BELOW the boundary.
        pt.append(_tracker_row(None, "Meridian Corp", "new job, no number yet",
                               [("VPO-N1", "2026-09-09")]))
    else:
        pt.append([None] * 8)
    pt.append([None] * 8)
    for lab in (L_ADMIN, L_OWNER, L_AWAIT):
        pt.append([None] * 5 + [lab])
    if footer_stray_leg:
        # A footer line with its label AND one stray character in a vendor-PO
        # column. One cell is not a job; treating it as one invents a vendor
        # leg on a phantom project, one click from create_project.
        pt.append([None] * 5 + ["Colour key updated 8/9", "x", None])
    elif stray_col_a:
        # A date stamp typed under the legend. Column A, nothing else. Enough,
        # once, to drag the boundary past the whole legend block.
        pt.append(["Updated 8/9/26"] + [None] * 7)
    else:
        pt.append([None] * 8)
    pt.append(["Invoice Number:", "PO#", "Invoice Date", "DUE Date",
               "Client", "Notes"])

    solid = {2: MAGENTA, 4: None}
    for row, argb in solid.items():
        if argb:
            pt.cell(row=row, column=6).fill = PatternFill(
                start_color=argb, end_color=argb, fill_type="solid")
    if keyed_bodyless_unknown:
        pt.cell(row=5, column=6).fill = PatternFill(
            start_color=UNKNOWN, end_color=UNKNOWN, fill_type="solid")
    if theme_row:
        # A colour from the top row of Excel's fill dropdown. openpyxl hands
        # back the RGB DESCRIPTOR here, not a string.
        pt.cell(row=3, column=6).fill = PatternFill(
            fill_type="solid", fgColor=Color(theme=4, tint=0.4))
    if default_fill_row:
        # solid with no explicit foreground -> openpyxl reports '00000000'
        pt.cell(row=4, column=6).fill = PatternFill(fill_type="solid")
    for row in (7, 8, 9):
        argb = {7: MAGENTA, 8: YELLOW, 9: CYAN}[row]
        pt.cell(row=row, column=6).fill = PatternFill(
            start_color=argb, end_color=argb, fill_type="solid")
    if footer_stray_leg:
        pt.cell(row=10, column=6).fill = PatternFill(
            start_color=MAGENTA, end_color=MAGENTA, fill_type="solid")
    if legend_unreadable:
        # The legend row is right there on the sheet; only its colour cannot be
        # read. Reporting that as "no legend row found" sends him looking for a
        # row that is not missing.
        pt.cell(row=7, column=6).fill = PatternFill(
            fill_type="solid", fgColor=Color(theme=4, tint=0.4))

    for n in ("Client Contacts", "Vendor Contacts"):
        wb.create_sheet(n).append(["x"])
    wb.save(path)


def _import_edge(nrm, crm, tmp, name, **kw):
    xl = tmp / f"{name}.xlsx"
    _build_edge_workbook(xl, **kw)
    out = tmp / f"store-{name}"
    out.mkdir()
    added = str(crm / "pipeline")
    did = added not in sys.path
    if did:
        sys.path.insert(0, added)
    err = None
    try:
        nrm.run(str(xl), str(out), force=False, mode="merge")
    except Exception as exc:                                  # noqa: BLE001
        err = f"{type(exc).__name__}: {exc}"
    finally:
        if did:
            try:
                sys.path.remove(added)
            except ValueError:
                pass

    def load(entity):
        p = out / f"{entity}.json"
        return json.loads(p.read_text()) if p.exists() else None
    return err, load


def _import(nrm, crm, tmp, name, **wbkw):
    """Build a fixture workbook, import it, return (error, loader)."""
    xl = tmp / f"{name}.xlsx"
    _build_workbook(xl, **wbkw)
    out = tmp / f"store-{name}"
    out.mkdir()
    added = str(crm / "pipeline")
    did_add = added not in sys.path
    if did_add:
        sys.path.insert(0, added)       # normalize imports merge as a sibling
    err = None
    try:
        nrm.run(str(xl), str(out), force=False, mode="merge")
    except Exception as exc:                                  # noqa: BLE001
        err = f"{type(exc).__name__}: {exc}"
    finally:
        if did_add:
            try:
                sys.path.remove(added)
            except ValueError:
                pass

    def load(entity):
        p = out / f"{entity}.json"
        return json.loads(p.read_text()) if p.exists() else None
    return err, load


def _merge_checks(r, crm, tmp):
    """Re-import behaviour for the tracker fields.

    Both cases here are UPGRADE-PATH bugs: they only appear on the second
    import, which is why asserting the REGENERATED set's source text was not
    enough. The screen the operator lands on is built from the merged store,
    not from the import."""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "_lt_merge", crm / "pipeline" / "merge.py")
    merge = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(merge)

    def store(name, projects, changelog=None):
        d = tmp / f"mg-{name}"
        d.mkdir(parents=True, exist_ok=True)
        for e in ("companies", "contacts", "shipments", "vendors",
                  "needs_review", "invoices"):
            (d / f"{e}.json").write_text("[]")
        (d / "companies.json").write_text(json.dumps(
            [{"company_id": "acme", "display_name": "Ace", "role": "customer",
              "archived": False}]))
        (d / "projects.json").write_text(json.dumps(projects))
        if changelog is not None:
            (d / "changelog.jsonl").write_text(changelog)
        return d

    def fresh(projects, unlinked):
        return {"companies.json": [{"company_id": "acme", "display_name": "Ace",
                                    "role": "customer", "archived": False}],
                "contacts.json": [], "shipments.json": [], "invoices.json": [],
                "vendors.json": [], "needs_review.json": [],
                "tracker_buckets.json": [], "tracker_unlinked.json": unlinked,
                "projects.json": projects}

    # ---- add-only mode still has to land the status -----------------------
    # changelog.jsonl is created lazily, on the first edit. A store that was
    # imported and never edited has data and NO changelog, which puts merge in
    # add-only mode: every existing record untouched. That meant tracker_status
    # never reached a single pre-existing project, so the operator upgraded,
    # imported, landed on the new default screen and read "No live projects
    # yet" -- permanently, on every subsequent import too.
    r.section("upgrading a store with no changelog still fills the tracker")
    d = store("addonly", [{"project_no": "4521", "company_id": "acme",
                           "status": "won", "revenue": 50000,
                           "archived": False}])
    merged, rep = merge.merge_all(
        fresh([{"project_no": "4521", "company_id": "acme", "status": "won",
                "revenue": 50000, "archived": False,
                "tracker_status": "action_admin", "tracker_row": 4,
                "open_orders_notes": "from the sheet"}], []), str(d))
    got = merged["projects.json"][0]
    r.check("add-only mode is what is under test here",
            "no changelog" in str(rep.get("note") or ""),
            f"got note {rep.get('note')!r} -- if this is not add-only mode the "
            f"checks below prove nothing")
    r.check("the status colour reaches an existing project anyway",
            got.get("tracker_status") == "action_admin",
            f"got {got.get('tracker_status')!r} -- the workbook owns this "
            f"field outright and nothing in the app can edit it, so there is "
            f"no operator edit for add-only mode to be protecting")
    r.check("and so does the row it came from", got.get("tracker_row") == 4)
    r.check("the operator's own revenue figure is still untouched",
            got.get("revenue") == 50000)
    r.check("the note is NOT force-refreshed in add-only mode",
            "open_orders_notes" not in got,
            "unlike the status, the note IS writable through the tools, so in "
            "add-only mode it stays the operator's")
    r.check("and the report says the status is the exception",
            "status" in str(rep.get("note") or "").lower(),
            f"got {rep.get('note')!r} -- a mode that says it refreshed nothing "
            f"while refreshing something is worse than either")

    # The other direction. In add-only mode a cleared colour does NOT retire:
    # "absent from the fresh record" is indistinguishable from "the tracker
    # matched nothing this run" (normalize only attaches these fields to a
    # project a tracker row matched), and popping on absence emptied the whole
    # Live screen off one import from an older copy of the workbook -- and took
    # tracker_row with it, which the importer can never re-derive. A stale
    # bucket is recoverable by importing the right workbook. A wiped one is not.
    d = store("addonly-retire", [{"project_no": "4521", "company_id": "acme",
                                  "status": "won", "archived": False,
                                  "tracker_status": "action_admin",
                                  "tracker_row": 4}])
    merged, _ = merge.merge_all(
        fresh([{"project_no": "4521", "company_id": "acme", "status": "won",
                "archived": False}], []), str(d))
    got = merged["projects.json"][0]
    r.check("add-only mode refreshes the status but never removes it",
            got.get("tracker_status") == "action_admin",
            f"got {got.get('tracker_status')!r} -- in add-only mode an absent "
            f"field means 'this workbook told us nothing', not 'retired'")
    r.check("and keeps the row reference the adopt flow wrote",
            got.get("tracker_row") == 4,
            f"got {got.get('tracker_row')!r} -- only saveAdoptTrackerRow ever "
            f"writes this for a numberless row; the importer cannot put it back")

    # With a changelog present the ordinary rule applies and a cleared colour
    # DOES retire -- so the leniency above is scoped to add-only mode, not a
    # blanket refusal to ever retire a card.
    d = store("retire-normal", [{"project_no": "4521", "company_id": "acme",
                                 "status": "won", "archived": False,
                                 "tracker_status": "action_admin",
                                 "tracker_row": 4}],
              changelog=json.dumps({"entity": "project", "key": "4521",
                                    "op": "update", "fields": ["revenue"]}) + "\n")
    merged, _ = merge.merge_all(
        fresh([{"project_no": "4521", "company_id": "acme", "status": "won",
                "archived": False}], []), str(d))
    got = merged["projects.json"][0]
    r.check("a colour cleared in Excel retires normally",
            not got.get("tracker_status"),
            f"got {got.get('tracker_status')!r} -- the row is off the tracker, "
            f"so a card for it is work the sheet says is finished")

    # ---- an adopted row must not come back --------------------------------
    r.section("a row already adopted is not offered for adoption again")
    log = json.dumps({"entity": "project", "key": "1419", "op": "create",
                      "fields": ["project_no"]}) + "\n"
    d = store("adopted", [{"project_no": "1419", "company_id": "acme",
                           "status": "won", "archived": False,
                           "tracker_status": "action_admin"}], changelog=log)
    merged, rep = merge.merge_all(
        fresh([], [{"sheet_row": 8, "reason": "no matching project",
                    "raw_key": "1419", "client": "Ironvale Supply",
                    "open_orders_notes": "keyed but unmatched", "legs": []},
                   {"sheet_row": 9, "reason": "no project number",
                    "client": "Meridian Corp",
                    "open_orders_notes": "still no number", "legs": []}]),
        str(d))
    left = [u.get("sheet_row") for u in merged["tracker_unlinked.json"]]
    r.check("the keyed row he adopted is gone from the list", left == [9],
            f"got rows {left} -- the sheet still has no CRM number on that "
            f"row, so it is regenerated every import; 'Add to CRM' on the "
            f"phantom fails with 'already exists' and the card only clears on "
            f"SUCCESS, so it is undismissable and returns forever")
    r.check("the one still unadopted stays", 9 in left)
    r.check("and the report names what it dropped",
            "1419" in str(rep.get("adopted")),
            f"got {rep.get('adopted')!r} -- a row vanishing off the screen "
            f"with no record is the same class of silence this fixes")

    # A NUMBERLESS row has no key, and nothing else about it is stable enough
    # to match on. An earlier attempt matched (tracker_row, note); it dropped a
    # live job whenever both notes were empty (the tuple degenerates to the row
    # number, which moves week to week), and it never fired anyway once the
    # operator edited the note in the adopt form -- which the form invites. So
    # the card stays. A duplicate card is visible and survivable; a dropped job
    # is neither.
    d = store("adopted-nokey", [{"project_no": "1500", "company_id": "acme",
                                 "status": "won", "archived": False,
                                 "tracker_row": 9,
                                 "open_orders_notes": "still no number"}],
              changelog=log)
    merged, rep = merge.merge_all(
        fresh([], [{"sheet_row": 9, "reason": "no project number",
                    "client": "Meridian Corp",
                    "open_orders_notes": "still no number", "legs": []}]),
        str(d))
    r.check("a numberless row is left showing rather than guessed at",
            [u.get("sheet_row") for u in merged["tracker_unlinked.json"]] == [9],
            f"got {merged['tracker_unlinked.json']!r} -- matching it on the "
            f"sheet row hides a live job the week that row is reused, and this "
            f"file calls that worse than a duplicate card")
    r.check("and nothing claims it was adopted", not rep.get("adopted"),
            f"got {rep.get('adopted')!r}")

    # the ".0" shape that has bitten this module before: raw_key is stored with
    # a bare str(), so a numeric key cell arrives as "1419.0"
    d = store("adopted-float", [{"project_no": "1419", "company_id": "acme",
                                 "status": "won", "archived": False}],
              changelog=log)
    merged, _ = merge.merge_all(
        fresh([], [{"sheet_row": 8, "reason": "no matching project",
                    # exactly the shape normalize writes for a NUMERIC key
                    # cell: the display string keeps the ".0", the parsed key
                    # does not. Matching on the display string re-admits the
                    # phantom; stripping ".0" from it would undo 0.1.28.
                    "raw_key": "1419.0", "parsed_keys": ["1419"],
                    "client": "Ironvale Supply",
                    "open_orders_notes": "keyed but unmatched", "legs": []}]),
        str(d))
    r.check("a .0-suffixed sheet key still matches the project it became",
            merged["tracker_unlinked.json"] == [],
            f"got {merged['tracker_unlinked.json']!r} -- the row shows as "
            f"'1419.0' and the project is '1419'; matching has to use the "
            f"parse both sides came through, not the display string")

    # and the report has to SAY what it took off the screen
    d = store("adopted-report", [{"project_no": "1419", "company_id": "acme",
                                  "status": "won", "archived": False}],
              changelog=log)
    merged, rep = merge.merge_all(
        fresh([], [{"sheet_row": 8, "reason": "no matching project",
                    "raw_key": "1419", "client": "Ironvale Supply",
                    "open_orders_notes": "keyed but unmatched", "legs": []}]),
        str(d))
    out = merge.format_report(rep)
    r.check("the printed report names the row it stopped offering",
            "1419" in out and "Not in the CRM yet" in out,
            f"got:\n{out}\n-- the drop site's two skip-ok markers both claim "
            f"it is recorded in the report, and that has to be true: a row "
            f"leaving the screen with nothing said is the silence this "
            f"module exists to end")


def run(server, crm_dir=None):
    r = Result("live-tracker/import", since="0.1.31")
    crm = Path(crm_dir) if crm_dir else \
        Path(__file__).resolve().parents[2] / "plugins/unrivaled-solutions/skills/crm"
    nrm = _load(crm, "normalize.py")
    if nrm is None:
        r.check("pipeline/normalize.py exists", False)
        return r

    # ---- the colour table itself --------------------------------------------
    # Asserted before the import runs, because every check below reads its
    # answers through it: if the constants are wrong the fixture is wrong too
    # and the whole module agrees with itself about nothing.
    r.section("the colour table")
    have = getattr(nrm, "BUCKET_BY_ARGB", None)
    if not isinstance(have, dict):
        r.check("BUCKET_BY_ARGB exists", False, "no colour table to decode with")
        return r
    r.check("three bucket colours, no more", len(have) == 3,
            f"got {sorted(have)} -- a fourth would group live work under a "
            f"heading the legend never names")
    for argb, key in ((MAGENTA, "action_admin"), (YELLOW, "action_owner"),
                      (CYAN, "awaiting_materials")):
        r.check(f"{argb} decodes to {key}", have.get(argb) == key,
                f"got {have.get(argb)!r}")
    nonb = getattr(nrm, "NON_BUCKET_ARGB", {})
    r.check("FF00FF00 is registered as explicitly NOT a bucket",
            GREEN in nonb,
            "without an entry it is an unknown colour and flags every import, "
            "which is how the review list becomes noise nobody reads")
    r.check("and it is not ALSO in the bucket table", GREEN not in have,
            "the second table's section headers would become live projects")
    r.check("the never-a-bucket entry says what the colour is",
            bool(str(nonb.get(GREEN, "")).strip()),
            "a bare colour code in a suppression list is unmaintainable")

    # A label read from the sheet must never be committed to source. On the
    # real workbook these name people and this repo is public.
    src = (crm / "pipeline" / "normalize.py").read_text()
    r.check("no bucket LABEL is hardcoded in the importer",
            "tracker_buckets.json" in src
            and not any(w in src for w in ("action_admin\":", "= \"Waiting")),
            "the labels are read from the legend at import; a hardcoded one "
            "both leaks a name and stops following a retitled bucket")

    # ---- the import, against a real workbook with real fills ----------------
    import openpyxl  # noqa: F401 -- see test_importer.py on why this is bare
    tmp = Path(tempfile.mkdtemp(prefix="crmlt-"))
    try:
        err, load = _import(nrm, crm, tmp, "full")
        r.check("the import runs to completion", err is None, err or "")
        if err is not None:
            return r

        buckets = load("tracker_buckets")
        unlinked = load("tracker_unlinked")
        projects = load("projects")
        shipments = load("shipments")
        review = load("needs_review")
        by_no = {str(p.get("project_no")): p for p in (projects or [])}

        r.check("tracker_buckets.json is written", isinstance(buckets, list),
                "the view falls back to raw keys and the screen names nobody")
        r.check("tracker_unlinked.json is written", isinstance(unlinked, list))
        if buckets is None or unlinked is None:
            return r

        # ---- the legend, found by structure -------------------------------
        r.section("the legend is anchored on the last keyed row, not on colour")
        labels = {b.get("key"): b.get("label") for b in buckets}
        r.check("the magenta bucket takes its name from the legend",
                labels.get("action_admin") == L_ADMIN,
                f"got {labels.get('action_admin')!r}")
        r.check("the yellow bucket takes its name from the legend",
                labels.get("action_owner") == L_OWNER,
                f"got {labels.get('action_owner')!r}")
        r.check("the cyan bucket takes its name from the legend",
                labels.get("awaiting_materials") == L_AWAIT,
                f"got {labels.get('awaiting_materials')!r}")
        r.check("exactly three buckets are produced", len(buckets) == 3,
                f"got {[b.get('key') for b in buckets]}")
        rows = {b.get("key"): b.get("legend_row") for b in buckets}
        r.check("each bucket records which legend row named it",
                rows == {"action_admin": 11, "action_owner": 12,
                         "awaiting_materials": 13},
                f"got {rows} -- expected the three rows BELOW the last keyed "
                f"row (9), not the coloured data rows above it")

        # The trap, stated as its own assertion: row 7 is cyan, above the
        # boundary, and carries text in the notes column. A colour-anchored
        # scan names the cyan bucket after it.
        r.check("a coloured DATA row above the boundary is not read as legend",
                labels.get("awaiting_materials") != "No number on this one yet",
                "row 8 is a real unkeyed job, cyan, with a note. Scanning for "
                "colour finds it first and the sheet's actual legend then gets "
                "imported as live work")

        # ...and the other half of the same bug: the footer must not become work.
        pnos = set(by_no)
        r.check("no legend row was imported as a project",
                not (pnos & {L_ADMIN, L_OWNER, L_AWAIT}),
                f"project keys were {sorted(pnos)}")
        notes_seen = {str(p.get("open_orders_notes")) for p in projects}
        r.check("no legend label ended up as a project's note",
                not (notes_seen & {L_ADMIN, L_OWNER, L_AWAIT}),
                "the five footer rows were swept into the live list once")
        r.check("the second table's header row is not a project",
                "CLIENT INVOICES BELOW" not in pnos and
                "CLIENT INVOICES BELOW" not in notes_seen)

        # ---- decoding a bucket off a fill ---------------------------------
        r.section("status comes off the fill colour")
        r.check("a magenta row is action_admin",
                by_no.get("5001", {}).get("tracker_status") == "action_admin",
                f"got {by_no.get('5001', {}).get('tracker_status')!r}")
        r.check("a yellow row is action_owner",
                by_no.get("5002", {}).get("tracker_status") == "action_owner",
                f"got {by_no.get('5002', {}).get('tracker_status')!r}")
        r.check("a cyan row is awaiting_materials",
                by_no.get("5003", {}).get("tracker_status") == "awaiting_materials",
                f"got {by_no.get('5003', {}).get('tracker_status')!r}")
        r.check("the row it came from is recorded",
                by_no.get("5001", {}).get("tracker_row") == 2,
                f"got {by_no.get('5001', {}).get('tracker_row')!r} -- without "
                f"it a review entry cannot be looked up in the sheet")

        # ---- an unknown colour is never guessed ----------------------------
        r.section("an unrecognised fill is flagged, not guessed")
        r.check("an unknown ARGB leaves the status unset",
                by_no.get("5004", {}).get("tracker_status") is None,
                f"got {by_no.get('5004', {}).get('tracker_status')!r} -- "
                f"nearest-colour matching files a job under the wrong person")
        unk = [x for x in review
               if x.get("type") == "tracker_unknown_status_colour"]
        r.check("exactly one unknown-colour entry is raised", len(unk) == 1,
                f"got {len(unk)}: {[x.get('sheet_row') for x in unk]}")
        if unk:
            r.check("it names the ARGB it could not decode",
                    UNKNOWN in str(unk[0].get("detail", "")),
                    f"got {unk[0].get('detail')!r}")
            r.check("and the row and sheet to look it up in",
                    unk[0].get("sheet_row") == 5
                    and unk[0].get("sheet") == "Project Tracker",
                    f"got row {unk[0].get('sheet_row')!r} "
                    f"sheet {unk[0].get('sheet')!r}")

        r.section("FF00FF00 is never a bucket and never a complaint")
        r.check("a green row gets no status",
                by_no.get("5005", {}).get("tracker_status") is None,
                f"got {by_no.get('5005', {}).get('tracker_status')!r}")
        r.check("and raises NO unknown-colour flag",
                all(x.get("sheet_row") != 6 for x in unk),
                "one flag per import per green row is what buries the real ones")
        r.check("green produced no bucket to group under",
                all(b.get("argb") != GREEN for b in buckets),
                f"got {[b.get('argb') for b in buckets]}")

        r.check("a complete legend raises no missing-legend flag",
                not [x for x in review
                     if x.get("type") == "tracker_legend_missing"],
                "this fixture has all three legend rows; a flag here means the "
                "scan is not finding them and every label check above is "
                "passing for some other reason")

        # ---- the note is a property of the row -----------------------------
        r.section("the note belongs to the project, not to each leg")
        r.check("a project with NO vendor legs still carries its note",
                by_no.get("5002", {}).get("open_orders_notes") == NOTE_NO_LEGS,
                f"got {by_no.get('5002', {}).get('open_orders_notes')!r} -- on "
                f"the shipment this note had nowhere to live at all")
        legs_5002 = [s for s in shipments if str(s.get("project_no")) == "5002"]
        r.check("and that project really has no legs", len(legs_5002) == 0,
                f"got {len(legs_5002)} -- the check above would be trivial")
        r.check("a project with five legs carries ONE note",
                by_no.get("5003", {}).get("open_orders_notes") == NOTE_FIVE_LEGS,
                f"got {by_no.get('5003', {}).get('open_orders_notes')!r}")
        legs_5003 = [s for s in shipments if str(s.get("project_no")) == "5003"]
        r.check("the five-leg row really has five legs", len(legs_5003) == 5,
                f"got {len(legs_5003)} -- if the U/V pair is dropped this is 4, "
                f"and the busiest live row loses a vendor with no error")
        r.check("the leg in column U/V is imported",
                any(s.get("vendor_po_raw") == "VPO-U8" for s in legs_5003),
                f"got {[s.get('vendor_po_raw') for s in legs_5003]} -- the old "
                f"bound stopped at S/T")
        r.check("the note is read from the notes column",
                by_no.get("5001", {}).get("open_orders_notes") == "On the bench",
                "this check was once labelled as covering the EMPTY-cell case "
                "while asserting a non-empty one, so the `if clean(cells[5])` "
                "branch had no coverage at all -- 5006 below is the real case")
        r.check("a row whose notes cell IS empty gets no note",
                by_no.get("5006", {}).get("open_orders_notes") in (None, ""),
                f"got {by_no.get('5006', {}).get('open_orders_notes')!r} -- an "
                f"empty cell must not write an empty string over anything")
        r.check("and still gets its status, which is the fill not the text",
                by_no.get("5006", {}).get("tracker_status") == "action_owner",
                f"got {by_no.get('5006', {}).get('tracker_status')!r}")

        # ---- unlinked rows: shown, never invented --------------------------
        r.section("a row that cannot be matched is kept whole, not keyed")
        by_row = {u.get("sheet_row"): u for u in unlinked}
        r.check("exactly two rows are unlinked", len(unlinked) == 2,
                f"got rows {sorted(by_row)} -- expected 8 (no number) and "
                f"9 (no matching project)")
        u7, u8 = by_row.get(8), by_row.get(9)
        r.check("the row with no project number is kept", u7 is not None)
        r.check("the row keyed to nothing in the deal log is kept",
                u8 is not None)
        if u7:
            r.check("it says WHY it is unlinked",
                    u7.get("reason") == "no project number",
                    f"got {u7.get('reason')!r}")
            r.check("its client survives", u7.get("client") == "Meridian Corp",
                    f"got {u7.get('client')!r}")
            r.check("its note survives in full",
                    u7.get("open_orders_notes") == "No number on this one yet",
                    f"got {u7.get('open_orders_notes')!r}")
            r.check("its status bucket survives",
                    u7.get("tracker_status") == "awaiting_materials",
                    f"got {u7.get('tracker_status')!r} -- it is cyan on the "
                    f"sheet and the screen groups it by that")
            r.check("its vendor legs survive",
                    [l.get("vendor_po_raw") for l in u7.get("legs", [])]
                    == ["VPO-M1"],
                    f"got {u7.get('legs')!r}")
        if u8:
            r.check("the unmatched-key row says why",
                    u8.get("reason") == "no matching project",
                    f"got {u8.get('reason')!r}")
            r.check("it carries the PARSED key merge matches on",
                    u8.get("parsed_keys") == ["5999"],
                    f"got {u8.get('parsed_keys')!r} -- without this the merge "
                    f"falls back to the display string, which for a numeric "
                    f"key cell is '5999.0' and matches no project")
            r.check("and keeps the key the sheet actually carried",
                    str(u8.get("raw_key")) == "5999",
                    f"got {u8.get('raw_key')!r} -- he needs to recognise the "
                    f"row when he gives it a real number")
        r.check("NO synthetic project key was minted for either",
                "5999" not in pnos and len(projects) == 6,
                f"project keys were {sorted(pnos)} -- an invented key silently "
                f"mis-attaches the next edit made against it")
        r.check("both rows are ALSO named in the review list",
                len([x for x in review if x.get("type") in
                     ("open_order_row_without_project",
                      "tracker_row_without_matching_project")]) == 2,
                "the screen shows them; the review list is what he reads when "
                "he is not on that screen")

        # ---- a bucket the legend never names -------------------------------
        # Only its NAME is missing. The projects still have to group under it,
        # and the importer has to say the name is missing rather than drop the
        # bucket -- a dropped bucket takes its live projects off the screen.
        r.section("a bucket with no legend row still groups")
        err2, load2 = _import(nrm, crm, tmp, "nolegend",
                              legend=(L_ADMIN, L_OWNER, None))
        r.check("the import survives a legend row with no text", err2 is None,
                err2 or "")
        if err2 is None:
            b2 = load2("tracker_buckets") or []
            k2 = {b.get("key"): b for b in b2}
            r.check("the unnamed bucket is still present",
                    "awaiting_materials" in k2,
                    f"got {sorted(k2)} -- dropping it takes its live projects "
                    f"off the screen entirely")
            r.check("with no invented name",
                    k2.get("awaiting_materials", {}).get("label") is None,
                    f"got {k2.get('awaiting_materials', {}).get('label')!r}")
            r.check("and the two named ones keep their names",
                    k2.get("action_admin", {}).get("label") == L_ADMIN
                    and k2.get("action_owner", {}).get("label") == L_OWNER)
            miss = [x for x in (load2("needs_review") or [])
                    if x.get("type") == "tracker_legend_missing"]
            r.check("the missing legend is reported once", len(miss) == 1,
                    f"got {len(miss)}")
            if miss:
                r.check("and names the colour that has no label",
                        CYAN in str(miss[0].get("detail", "")),
                        f"got {miss[0].get('detail')!r}")
            p2 = {str(p.get("project_no")): p for p in (load2("projects") or [])}
            r.check("its projects still carry the bucket key",
                    p2.get("5003", {}).get("tracker_status")
                    == "awaiting_materials",
                    f"got {p2.get('5003', {}).get('tracker_status')!r}")

        # ---- the boundary, in the direction the trap row does not cover ----
        #
        # Row 7 above proves a coloured DATA row cannot be mistaken for the
        # legend. These prove the other two things the boundary must not do:
        # swallow a real row that happens to sit below it, and be moved by
        # something that is not a data row at all.
        r.section("the boundary skips footer rows without eating data")
        err3, load3 = _import_edge(nrm, crm, tmp, "trailing",
                                   trailing_unkeyed=True)
        r.check("the import survives a trailing unkeyed row", err3 is None,
                err3 or "")
        if err3 is None:
            u3 = load3("tracker_unlinked") or []
            rv3 = load3("needs_review") or []
            r.check("a real unkeyed job BELOW the boundary is still rescued",
                    [x.get("sheet_row") for x in u3] == [5],
                    f"got rows {[x.get('sheet_row') for x in u3]} -- appending "
                    f"a job at the bottom before it has a number is the normal "
                    f"way a row gets added, and it was being dropped in total "
                    f"silence: no review entry, no card, note and legs gone")
            r.check("and it is flagged as well as shown",
                    any(x.get("type") == "open_order_row_without_project"
                        and x.get("sheet_row") == 5 for x in rv3),
                    "before this feature existed the row was flagged wherever "
                    "it sat; the boundary made the importer quieter about real "
                    "data than it used to be")
            r.check("its legs come with it",
                    bool(u3) and [l.get("vendor_po_raw")
                                  for l in u3[0].get("legs", [])] == ["VPO-N1"],
                    f"got {u3[:1]}")
            r.check("the legend rows are still NOT rescued as work",
                    not ({L_ADMIN, L_OWNER, L_AWAIT} &
                         {x.get("open_orders_notes") for x in u3}),
                    "a footer row carries nothing but a notes cell; that is "
                    "what makes it a footer row rather than its position")
            b3 = {b.get("key"): b.get("label") for b in (load3("tracker_buckets") or [])}
            r.check("and the legend is still read",
                    b3.get("action_admin") == L_ADMIN, f"got {b3}")

        err4, load4 = _import_edge(nrm, crm, tmp, "stray", stray_col_a=True)
        r.check("the import survives a stray value under the legend",
                err4 is None, err4 or "")
        if err4 is None:
            b4 = {b.get("key"): b.get("label") for b in (load4("tracker_buckets") or [])}
            r.check("a column-A-only row below the legend does not move the "
                    "boundary",
                    b4 == {"action_admin": L_ADMIN, "action_owner": L_OWNER,
                           "awaiting_materials": L_AWAIT},
                    f"got {b4} -- a date stamp, a TOTAL, or the text of a "
                    f"merged footer comment all land in column A, and any one "
                    f"of them used to blank all three bucket names")
            u4 = load4("tracker_unlinked") or []
            r.check("and does not turn the legend into adoptable cards",
                    not ({L_ADMIN, L_OWNER, L_AWAIT} &
                         {x.get("open_orders_notes") for x in u4}),
                    f"got {[x.get('open_orders_notes') for x in u4]} -- these "
                    f"render with an 'Add to CRM' button on the operator's own "
                    f"legend text")
            r.check("no missing-legend flags either",
                    not [x for x in (load4("needs_review") or [])
                         if x.get("type") == "tracker_legend_missing"])

        # ---- a fill that is there but cannot be read -----------------------
        r.section("an unreadable fill is flagged, not treated as no fill")
        err5, load5 = _import_edge(nrm, crm, tmp, "theme", theme_row=True,
                                   default_fill_row=True)
        r.check("the import survives a theme fill", err5 is None, err5 or "")
        if err5 is None:
            p5 = {str(p.get("project_no")): p for p in (load5("projects") or [])}
            rv5 = [x for x in (load5("needs_review") or [])
                   if x.get("type") == "tracker_unknown_status_colour"]
            r.check("a theme-coloured row gets no guessed bucket",
                    p5.get("6002", {}).get("tracker_status") is None,
                    f"got {p5.get('6002', {}).get('tracker_status')!r}")
            r.check("but it IS reported, naming the row",
                    any(x.get("sheet_row") == 3 for x in rv5),
                    f"got {[x.get('sheet_row') for x in rv5]} -- a theme colour "
                    f"made .rgb return a descriptor rather than a string, the "
                    f"row was dropped from the fills map, and the project then "
                    f"vanished off the Live screen with nothing said anywhere")
            r.check("the entry says the colour could not be read, not that it "
                    "was unrecognised",
                    any("theme" in str(x.get("detail", "")) for x in rv5),
                    f"got {[x.get('detail') for x in rv5]}")
            r.check("a solid fill with no foreground colour is NOT reported",
                    not any(x.get("sheet_row") == 4 for x in rv5),
                    "'00000000' is the default, i.e. no colour -- reporting it "
                    "puts an entry in the review list on every single import")
            r.check("and gets no bucket either",
                    p5.get("6003", {}).get("tracker_status") is None,
                    f"got {p5.get('6003', {}).get('tracker_status')!r}")

        err6, load6 = _import_edge(nrm, crm, tmp, "bodyless",
                                   keyed_bodyless_unknown=True)
        r.check("the import survives a keyed row with nothing else on it",
                err6 is None, err6 or "")
        if err6 is None:
            rv6 = [x for x in (load6("needs_review") or [])
                   if x.get("type") == "tracker_unknown_status_colour"]
            r.check("a keyed row below the boundary still gets its colour flag",
                    any(x.get("sheet_row") == 5 for x in rv6),
                    f"got rows {[x.get('sheet_row') for x in rv6]} -- a row "
                    f"with a key is a data row wherever it sits; tying the flag "
                    f"to the boundary lost the status AND any mention of it")

        err7, load7 = _import_edge(nrm, crm, tmp, "strayleg",
                                   footer_stray_leg=True)
        r.check("the import survives a footer line with a stray cell",
                err7 is None, err7 or "")
        if err7 is None:
            u7b = load7("tracker_unlinked") or []
            r.check("one stray cell in a footer line is not a vendor leg",
                    not any("Colour key updated" in str(x.get("open_orders_notes"))
                            for x in u7b),
                    f"got {u7b!r} -- it renders as a phantom project carrying "
                    f"an invented leg, with an 'Add to CRM' button on it")
            b7 = {b.get("key"): b.get("label") for b in (load7("tracker_buckets") or [])}
            r.check("and it does not move the boundary either",
                    b7.get("action_admin") == L_ADMIN, f"got {b7}")

        err8, load8 = _import_edge(nrm, crm, tmp, "legendfill",
                                   legend_unreadable=True)
        r.check("the import survives an unreadable legend fill", err8 is None,
                err8 or "")
        if err8 is None:
            rv8 = load8("needs_review") or []
            unread = [x for x in rv8
                      if x.get("type") == "tracker_legend_unreadable_colour"]
            r.check("an unreadable fill on a legend row says so",
                    len(unread) == 1,
                    f"got {[x.get('type') for x in rv8]} -- otherwise the row "
                    f"falls through to 'no legend row found for the FFFF00FF "
                    f"bucket', pointing him at a row that is right there")
            if unread:
                r.check("and names the legend text it could not colour-match",
                        L_ADMIN in str(unread[0].get("detail", "")),
                        f"got {unread[0].get('detail')!r}")

        # ---- the regeneration interlock ------------------------------------
        # Both files are the importer's reading of THIS run. Merging them by key
        # would resurrect a row he has already adopted into the CRM.
        r.section("both tracker files are regenerated, never merged")
        mg = (crm / "pipeline" / "merge.py").read_text()
        regen = mg.split("REGENERATED", 1)[-1].split("\n\n", 1)[0]
        r.check("tracker_unlinked.json is regenerated wholesale",
                "tracker_unlinked.json" in regen,
                "merged by key, a row adopted into the CRM comes back as an "
                "unadopted card on the next import")
        r.check("tracker_buckets.json is regenerated wholesale",
                "tracker_buckets.json" in regen,
                "a bucket he renamed in the sheet would keep its old label")

        # ---- re-import, where the tracker meets the operator's own work ----
        _merge_checks(r, crm, tmp)

        # ---- the fields have to be writable through the server -------------
        r.section("the new project fields are accepted by the write path")
        pf = getattr(server, "PROJECT_FIELDS", None) if server else None
        if pf is None:
            src_srv = (crm / "mcp" / "server.py").read_text()
            pf = set()
            for f in ("tracker_status", "open_orders_notes", "tracker_row"):
                if f'"{f}"' in src_srv:
                    pf.add(f)
        for f in ("tracker_status", "open_orders_notes", "tracker_row"):
            r.check(f"{f} is a writable project field", f in pf,
                    "adoption writes it through create_project; a field the "
                    "validator does not know is dropped on the floor")

        # ---- tracker_status is an enum, and every sibling enum is checked --
        #
        # It was the fourth state field on a project and the only one with no
        # validation. A value no bucket knows about left the project counted in
        # the header's "N active", rendered in no section, and still listed in
        # the sidebar -- a live job off the daily board with nothing said.
        r.section("tracker_status is validated like every other state field")
        keys = getattr(nrm, "TRACKER_BUCKET_KEYS", None) \
            or set(getattr(nrm, "BUCKET_BY_ARGB", {}).values())
        srv_keys = getattr(server, "TRACKER_STATUSES", None) if server else None
        if srv_keys is None:
            # The mutation runner calls run(None, crm_dir) -- there is no
            # imported server object there, so read the constant out of the
            # source. Without this the module fails at BASELINE and every
            # mutant result against server.py is a false positive.
            m = re.search(r"^TRACKER_STATUSES\s*=\s*\{([^}]*)\}",
                          (crm / "mcp" / "server.py").read_text(), re.M)
            if m:
                srv_keys = set(re.findall(r'"([^"]+)"', m.group(1)))
        if srv_keys is None:
            r.check("the server knows the bucket keys", False,
                    "no TRACKER_STATUSES -- nothing constrains the field")
        else:
            r.check("the server's bucket keys match the importer's",
                    set(srv_keys) == set(keys),
                    f"server {sorted(srv_keys)} vs importer {sorted(keys)} -- a "
                    f"key accepted by one and unknown to the other renders a "
                    f"live project into no section at all")
        # Load the server from THIS crm dir rather than using the one passed
        # in: the mutation runner calls run(None, crm_dir), so gating on the
        # argument meant every mutant of the validation itself survived while
        # only the constant was ever checked.
        from lib.harness import (Store, company, invoice, load_server,
                                 project, shipment)
        srv = server
        if srv is None:
            try:
                srv = load_server(str(crm))
            except Exception as exc:                          # noqa: BLE001
                r.check("the server module imports", False,
                        f"{type(exc).__name__}: {exc}")
        if srv is not None:
            st = Store(srv)
            st.reset(companies=[company()], projects=[project("4521")])
            ok = st.call("update_project", project_no="4521",
                         fields={"tracker_status": "action_admin"})
            r.check("a real bucket key is accepted", ok.get("ok") is True,
                    f"got {ok}")
            for bad in ("done", " action_admin", "Action Admin",
                        "awaiting_material", 5):
                res = st.call("update_project", project_no="4521",
                              fields={"tracker_status": bad})
                r.check(f"tracker_status {bad!r} is refused",
                        res.get("ok") is False and "_raised" not in res,
                        f"got {res} -- accepted, so the project is counted as "
                        f"live and rendered in no bucket")
            # ---- moving a project must carry its legs ------------------
            #
            # rename_project cascades a number change and reassign_shipment
            # carries a leg to its new project's customer, but update_project
            # wrote company_id onto the project alone. The Live Tracker matches
            # legs on company AND number -- two customers can hold one number,
            # and one customer's leg must never surface on another's card -- so
            # after a move the card read "No vendor legs" and every lateness
            # flag on that job stopped firing, on the screen built to show them.
            st.reset(companies=[company(), company("mer", "Meridian Corp")],
                     projects=[project("4521", "acme",
                                       tracker_status="action_admin")],
                     shipments=[shipment("4521-L1", "4521", "acme"),
                                shipment("4521-L2", "4521", "acme")],
                     invoices=[invoice("7001", "acme", project_no="4521")])
            mv = st.call("update_project", project_no="4521",
                         fields={"company_id": "mer",
                                 "company_name": "Meridian Corp"})
            r.check("moving a project to another company succeeds",
                    mv.get("ok") is True, f"got {mv}")
            r.check("and reports how many legs went with it",
                    mv.get("shipments_moved") == 2,
                    f"got {mv.get('shipments_moved')!r}")
            moved = [x for x in (st.read("shipments") or [])
                     if x.get("company_id") == "mer"]
            r.check("both legs are filed under the new company",
                    len(moved) == 2,
                    f"got {[(x.get('shipment_id'), x.get('company_id')) for x in (st.read('shipments') or [])]}")
            r.check("the leg's client name follows too",
                    all(x.get("client_name") == "Meridian Corp" for x in moved),
                    f"got {[x.get('client_name') for x in moved]}")
            r.check("its invoice moves as well",
                    mv.get("invoices_moved") == 1
                    and all(i.get("company_id") == "mer"
                            for i in (st.read("invoices") or [])),
                    f"got {st.read('invoices')!r}")

            # a leg belonging to a DIFFERENT company that happens to share the
            # number must NOT be dragged along
            st.reset(companies=[company(), company("mer", "Meridian Corp"),
                                company("nor", "Northgate Tooling")],
                     projects=[project("4521", "acme")],
                     shipments=[shipment("4521-L1", "4521", "acme"),
                                shipment("x-4521-L1", "4521", "nor")])
            st.call("update_project", project_no="4521",
                    fields={"company_id": "mer"})
            other = [x for x in (st.read("shipments") or [])
                     if x.get("shipment_id") == "x-4521-L1"]
            r.check("another company's leg on the same number is left alone",
                    bool(other) and other[0].get("company_id") == "nor",
                    f"got {other!r} -- the same number can belong to two "
                    f"customers, which is why the match is on both")

            # an ordinary edit that does not move the project touches nothing
            st.reset(companies=[company()],
                     projects=[project("4521", "acme")],
                     shipments=[shipment("4521-L1", "4521", "acme")])
            plain = st.call("update_project", project_no="4521",
                            fields={"notes": "just a note"})
            r.check("an edit that does not move the project reports no move",
                    plain.get("ok") is True
                    and "shipments_moved" not in plain,
                    f"got {plain}")

            st.reset(companies=[company()], projects=[project("4521")])
            res = st.call("update_project", project_no="4521",
                          fields={"tracker_status": None})
            r.check("clearing it back to unset is allowed",
                    res.get("ok") is True,
                    f"got {res} -- a row whose colour was removed in Excel has "
                    f"to be able to retire")
            shutil.rmtree(st.path, ignore_errors=True)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    return r
